"""
Belcris Inventory Bot — v3.0
Flask webhook mode (matches original deployment architecture).

Full command set matching the original cheat sheet:

  Inventory:
    /refresh           — Reload all data from Google Drive
    /summary           — Inventory snapshot (total stock, top items, expiry alerts)
    /search <kw>       — Search by keyword
    /check <code>      — Look up by SAP item code
    /category [name]   — Browse by product category
    /expiring [days]   — Items expiring in N days (default 30)
    /low [threshold]   — Items below stock level
    /warehouse <name>  — Items in a warehouse
    /top               — Top 20 items by quantity

  Accounts Receivable:
    /arsummary         — AR executive summary with aging buckets
    /ar <client>       — Outstanding balance for a client
    /client <client>   — Alias for /ar
    /aging             — AR aging summary
    /overdue [days]    — Top overdue clients (optional days threshold)
    /area <name>       — Receivables grouped by area/agent
    /agent <name>      — Clients under a specific agent
    /arsearch <kw>     — Search clients by name in AR
    /arrefresh         — Force-refresh AR data only

  Accounts Payable:
    /apsummary         — AP summary with aging buckets + top vendors
    /ap_summary        — Alias for /apsummary
    /ap <vendor>       — Unreleased payments for a vendor
    /vendor <vendor>   — Alias for /ap
    /apaging           — AP aging by bucket
    /apoverdue         — Vendors overdue 61+ days
    /aptop [n]         — Top N vendors by amount (default 10)
    /due_today         — Payments due this week
    /aprefresh         — Force-refresh AP data only
"""

import os
import io
import re
import json
import logging
import asyncio
import difflib
import threading
from datetime import datetime, timezone, timedelta, date
from email.utils import parsedate_to_datetime
from functools import wraps

import requests
import openpyxl
try:
    import pymysql
    import pymysql.cursors
    PYMYSQL_AVAILABLE = True
except ImportError:
    PYMYSQL_AVAILABLE = False
from flask import Flask, request as flask_request
from telegram import Update, BotCommand, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
from telegram.constants import ParseMode

# ──────────────────────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ["BOT_TOKEN"]
PHT = timezone(timedelta(hours=8))
PORT = int(os.environ.get("PORT", 8080))
WEBHOOK_HOST = os.environ.get("RAILWAY_STATIC_URL", "").strip()

# Google Drive file IDs
INVENTORY_FILE_ID = os.environ.get("INVENTORY_FILE_ID", "1_qBGc6JV2OGoeISDFlPftUs7aE5nVKNy")
AR_FILE_ID        = os.environ.get("AR_FILE_ID",        "1nmr-7YLCZe2dPXkAplseUxMGrlhaWlTS")
AP_FILE_ID        = os.environ.get("AP_FILE_ID",        "1ejtfkY-Y72LgXRPdLq5Z-cQZk3uD8IwO")

# Ham Portal DB (read-only sales queries)
PORTAL_DB_URL = os.environ.get("PORTAL_DB_URL", "")  # mysql://user:pass@host:port/db?ssl=...

ON_HOLD_WAREHOUSES = {
    "Mets Logistics Inc - CARMONA",
    "Production Warehouse",
    "Component Warehouse",
    "Christmas Hams Supplies",
    "Intensive Care Unit (ICU) Items",  # WCS08A — excluded from all results
}

REFRESH_INTERVAL_SECONDS = 30 * 60  # 30 minutes
LOW_STOCK_THRESHOLD = 100

# Product category keywords (order matters — first match wins)
CATEGORY_KEYWORDS = {
    "HAM":        ["ham"],
    "BACON":      ["bacon"],
    "TOCINO":     ["tocino"],
    "LONGGANISA": ["longganisa", "longanisa"],
    "HOTDOG":     ["hotdog", "hot dog", "frankfurter"],
    "SAUSAGE":    ["sausage"],
    "BUNDLE":     ["bundle", "pack"],
    "CORNED":     ["corned"],
    "LUNCHEON":   ["luncheon"],
    "LIVER":      ["liver"],
    "NUGGETS":    ["nugget"],
    "SISIG":      ["sisig"],
    "TAPA":       ["tapa"],
    "EMBUTIDO":   ["embutido"],
}

# ──────────────────────────────────────────────────────────────────────────────
# Access Control (v4.0)
# SQLite-backed: user log (passive collection), allowlist (registered users),
# and group whitelist (approved group chats are fully open for all members —
# e.g. the internal team group where the boss is present, no action needed).
# Files live next to bot.py (Railway ephemeral disk is fine — state is rebuilt
# from the log; the log is re-collected passively on every interaction).
# Env vars:
#   ADMIN_IDS            — comma-separated Telegram user IDs with full admin
#                          rights (e.g. "123456789,987654321"). Required for
#                          admin commands.
#   ACCESS_MODE          — "off" (no blocking, log only) | "soft" (warn
#                          unregistered, still allow) | "hard" (block
#                          unregistered from data cmds). Default: "off".
#   ALLOWED_GROUP_CHAT_IDS — comma-separated group Chat IDs (negative numbers)
#                          that are fully open for all members regardless of
#                          ACCESS_MODE. Recommended for the internal team chat.
#   ACCESS_DB_PATH       — optional override for the SQLite DB path.
# ──────────────────────────────────────────────────────────────────────────────
import sqlite3
import threading

_DB_PATH = os.environ.get("ACCESS_DB_PATH", os.path.join(os.path.dirname(os.path.abspath(__file__)), "access_control.db"))
ADMIN_IDS_ENV = os.environ.get("ADMIN_IDS", "").strip()
ADMIN_IDS: set[int] = set()
if ADMIN_IDS_ENV:
    for raw in ADMIN_IDS_ENV.split(","):
        raw = raw.strip()
        if raw.isdigit():
            ADMIN_IDS.add(int(raw))

ACCESS_MODE: str = os.environ.get("ACCESS_MODE", "off").strip().lower()
if ACCESS_MODE not in ("off", "soft", "hard"):
    ACCESS_MODE = "off"

ALLOWED_GROUP_IDS_ENV = os.environ.get("ALLOWED_GROUP_CHAT_IDS", "").strip()
ALLOWED_GROUP_CHAT_IDS: set[int] = set()
for raw in ALLOWED_GROUP_IDS_ENV.split(","):
    raw = raw.strip()
    if raw.lstrip("-").isdigit():
        ALLOWED_GROUP_CHAT_IDS.add(int(raw))

_db_lock = threading.Lock()


def _access_db() -> sqlite3.Connection:
    conn = sqlite3.connect(_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_access_tables(conn: sqlite3.Connection):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS seen_users (
            telegram_user_id INTEGER PRIMARY KEY,
            username TEXT,
            full_name TEXT,
            first_seen TEXT NOT NULL DEFAULT (datetime('now')),
            last_seen TEXT NOT NULL DEFAULT (datetime('now')),
            message_count INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS registered_users (
            telegram_user_id INTEGER PRIMARY KEY,
            registered_by INTEGER NOT NULL,
            registered_at TEXT NOT NULL DEFAULT (datetime('now')),
            note TEXT
        );
        CREATE TABLE IF NOT EXISTS seen_groups (
            telegram_chat_id INTEGER PRIMARY KEY,
            chat_type TEXT,
            chat_title TEXT,
            first_seen TEXT NOT NULL DEFAULT (datetime('now')),
            last_seen TEXT NOT NULL DEFAULT (datetime('now')),
            interaction_count INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS allowed_groups (
            telegram_chat_id INTEGER PRIMARY KEY,
            allowed_by INTEGER NOT NULL,
            allowed_at TEXT NOT NULL DEFAULT (datetime('now')),
            note TEXT
        );
        CREATE TABLE IF NOT EXISTS group_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_chat_id INTEGER NOT NULL,
            telegram_user_id INTEGER NOT NULL,
            full_name TEXT,
            message_text TEXT,
            timestamp TEXT NOT NULL DEFAULT (datetime('now'))
        );
    """)


def log_user(user_id: int, username: str | None, full_name: str | None):
    """Passively log every user who interacts with the bot (fire-and-forget)."""
    if not user_id:
        return
    try:
        with _db_lock:
            conn = _access_db()
            _ensure_access_tables(conn)
            now = datetime.now(PHT).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                "INSERT INTO seen_users (telegram_user_id, username, full_name, last_seen, message_count) "
                "VALUES (?, ?, ?, ?, 1) "
                "ON CONFLICT(telegram_user_id) DO UPDATE SET "
                "username = excluded.username, full_name = excluded.full_name, "
                "last_seen = excluded.last_seen, message_count = message_count + 1",
                (user_id, username, full_name, now),
            )
            conn.commit()
            conn.close()
    except Exception as e:
        logger.warning(f"Failed to log user {user_id}: {e}")


def log_chat(chat_id: int | None, chat_type: str | None, chat_title: str | None):
    """Passively log every chat (group/supergroup/private) the bot is used in."""
    if not chat_id:
        return
    try:
        with _db_lock:
            conn = _access_db()
            _ensure_access_tables(conn)
            now = datetime.now(PHT).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                "INSERT INTO seen_groups (telegram_chat_id, chat_type, chat_title, last_seen, interaction_count) "
                "VALUES (?, ?, ?, ?, 1) "
                "ON CONFLICT(telegram_chat_id) DO UPDATE SET "
                "chat_type = excluded.chat_type, chat_title = excluded.chat_title, "
                "last_seen = excluded.last_seen, interaction_count = interaction_count + 1",
                (chat_id, chat_type, chat_title, now),
            )
            conn.commit()
            conn.close()
    except Exception as e:
        logger.warning(f"Failed to log chat {chat_id}: {e}")


def log_group_message(chat_id: int, user_id: int, full_name: str | None, text: str):
    """Log group messages for AI summarization (only for whitelisted groups)."""
    if not chat_id or not text or chat_id >= 0:
        return
    try:
        with _db_lock:
            conn = _access_db()
            _ensure_access_tables(conn)
            now = datetime.now(PHT).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                "INSERT INTO group_messages (telegram_chat_id, telegram_user_id, full_name, message_text, timestamp) "
                "VALUES (?, ?, ?, ?, ?)",
                (chat_id, user_id, full_name, text, now),
            )
            # Keep only last 500 messages per chat to save space
            conn.execute(
                "DELETE FROM group_messages WHERE id IN (SELECT id FROM group_messages WHERE telegram_chat_id = ? ORDER BY id DESC LIMIT -1 OFFSET 500)",
                (chat_id,),
            )
            conn.commit()
            conn.close()
    except Exception as e:
        logger.warning(f"Failed to log group message: {e}")


def is_whitelisted_group(chat_id: int | None) -> bool:
    """True if the chat is an admin-approved group (fully open for all members)."""
    if not chat_id or chat_id >= 0:
        return False
    if chat_id in ALLOWED_GROUP_CHAT_IDS:
        return True
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        row = conn.execute(
            "SELECT 1 FROM allowed_groups WHERE telegram_chat_id = ?", (chat_id,)
        ).fetchone()
        conn.close()
    return row is not None


def allow_group(chat_id: int, allowed_by: int, note: str | None = None) -> bool:
    """Whitelist a group chat. Returns False if already allowed."""
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        existing = conn.execute(
            "SELECT 1 FROM allowed_groups WHERE telegram_chat_id = ?", (chat_id,)
        ).fetchone()
        if existing:
            conn.close()
            return False
        conn.execute(
            "INSERT INTO allowed_groups (telegram_chat_id, allowed_by, note) VALUES (?, ?, ?)",
            (chat_id, allowed_by, note),
        )
        conn.commit()
        conn.close()
    return True


def unallow_group(chat_id: int) -> bool:
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        cur = conn.execute(
            "DELETE FROM allowed_groups WHERE telegram_chat_id = ?", (chat_id,)
        )
        conn.commit()
        conn.close()
    return cur.rowcount > 0


def list_allowed_groups() -> list[sqlite3.Row]:
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        rows = conn.execute(
            "SELECT a.telegram_chat_id, a.allowed_by, a.allowed_at, a.note, "
            "       COALESCE(g.chat_title, '') AS chat_title, COALESCE(g.last_seen, '') AS last_seen "
            "FROM allowed_groups a "
            "LEFT JOIN seen_groups g ON g.telegram_chat_id = a.telegram_chat_id "
            "ORDER BY a.allowed_at DESC"
        ).fetchall()
        conn.close()
    return rows


def list_seen_groups(limit: int = 500) -> list[sqlite3.Row]:
    """All group chats the bot has been used in, most recent first."""
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        rows = conn.execute(
            "SELECT telegram_chat_id, chat_type, chat_title, first_seen, last_seen, interaction_count "
            "FROM seen_groups ORDER BY last_seen DESC LIMIT ?",
            (limit,),
        ).fetchall()
        conn.close()
    return rows


def is_registered(user_id: int) -> bool:
    if not user_id:
        return False
    if ACCESS_MODE == "off":
        return True
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        row = conn.execute("SELECT 1 FROM registered_users WHERE telegram_user_id = ?", (user_id,)).fetchone()
        conn.close()
    return row is not None


def register_user(user_id: int, registered_by: int, note: str | None = None) -> bool:
    """Register a user. Returns False if already registered."""
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        existing = conn.execute(
            "SELECT 1 FROM registered_users WHERE telegram_user_id = ?", (user_id,)
        ).fetchone()
        if existing:
            conn.close()
            return False
        conn.execute(
            "INSERT INTO registered_users (telegram_user_id, registered_by, note) VALUES (?, ?, ?)",
            (user_id, registered_by, note),
        )
        conn.commit()
        conn.close()
    return True


def unregister_user(user_id: int) -> bool:
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        cur = conn.execute("DELETE FROM registered_users WHERE telegram_user_id = ?", (user_id,))
        conn.commit()
        conn.close()
    return cur.rowcount > 0


def list_registered_users() -> list[sqlite3.Row]:
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        rows = conn.execute(
            "SELECT r.telegram_user_id, r.registered_by, r.registered_at, r.note, "
            "       COALESCE(s.username, '') AS username, COALESCE(s.full_name, '') AS full_name, "
            "       COALESCE(s.last_seen, '') AS last_seen "
            "FROM registered_users r "
            "LEFT JOIN seen_users s ON s.telegram_user_id = r.telegram_user_id "
            "ORDER BY r.registered_at DESC"
        ).fetchall()
        conn.close()
    return rows


def list_seen_users(limit: int = 500) -> list[sqlite3.Row]:
    """All users who have ever interacted with the bot, ordered by most recent."""
    with _db_lock:
        conn = _access_db()
        _ensure_access_tables(conn)
        rows = conn.execute(
            "SELECT telegram_user_id, username, full_name, first_seen, last_seen, message_count "
            "FROM seen_users ORDER BY last_seen DESC LIMIT ?",
            (limit,),
        ).fetchall()
        conn.close()
    return rows


def _is_admin(user_id: int | None) -> bool:
    return user_id in ADMIN_IDS


async def _access_gate(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Log the user/chat and enforce soft/hard blocking. Returns False if access denied.
    Whitelisted group chats (ALLOWED_GROUP_CHAT_IDS or /allowgroup) are fully open
    for all members regardless of ACCESS_MODE."""
    user = update.effective_user
    chat = update.effective_chat
    # Passively log the chat itself (builds the group list for /whitelistgroup)
    log_chat(chat.id if chat else None, chat.type if chat else None, chat.title if chat else None)
    if user is None:
        return True
    log_user(user.id, user.username, user.full_name)
    # Whitelisted groups are fully open for every member — no action needed.
    if chat is not None and is_whitelisted_group(chat.id):
        return True
    # Admins and registered users always pass; off mode always passes.
    if _is_admin(user.id) or is_registered(user.id):
        return True
    if ACCESS_MODE == "off":
        return True
    if ACCESS_MODE == "hard":
        await _send_blocked_notice(update.message, str(user.id))
        return False
    # soft mode
    await _send_blocked_notice(update.message, str(user.id))
    return True


async def _send_blocked_notice(update_or_query, username_hint: str):
    """Notify a user they are not registered. Works for both Message and CallbackQuery."""
    mode_label = "registered" if ACCESS_MODE == "hard" else "asked to register"
    text = (
        "🔒 *Access Notice*\n\n"
        "This bot is for authorized Belcris staff only.\n"
        "You have not been registered yet.\n\n"
        "Send `/myid` and share your User ID with the admin.\n"
        "The admin will register you, then you'll have full access.\n"
        f"_(Your ID: {username_hint})_"
    )
    try:
        if hasattr(update_or_query, "message") and update_or_query.message is not None:
            await update_or_query.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
        else:
            q = update_or_query
            try:
                await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
            except Exception:
                await q.answer(text, show_alert=True)
    except Exception:
        try:
            await update_or_query.message.reply_text(text)
        except Exception:
            pass


# ──────────────────────────────────────────────────────────────────────────────
# In-memory data store
# ──────────────────────────────────────────────────────────────────────────────
class DataStore:
    def __init__(self):
        self.inventory: list[dict] = []
        self.component_inventory: list[dict] = []  # Component Warehouse rows (PRD use)
        self.ar_rows: list[dict] = []
        self.ap_rows: list[dict] = []
        self.last_refresh: datetime | None = None
        # Source file timestamps (from Google Drive Last-Modified headers)
        self.inventory_source_ts: str = ""
        self.ar_source_ts: str = ""
        self.ap_source_ts: str = ""
        self._lock = threading.Lock()

store = DataStore()

# ──────────────────────────────────────────────────────────────────────────────
# Google Drive helpers
# ──────────────────────────────────────────────────────────────────────────────
def get_drive_file_modified_ts(file_id: str, is_sheets: bool = False) -> str:
    """
    Fetch the Last-Modified timestamp of a Google Drive file.
    Returns a formatted PHT string like '06/04/2026 02:52 PM' or '' on failure.
    Works for both regular Drive files and Google Sheets.
    Uses drive.usercontent.google.com which exposes Last-Modified for all file types.
    """
    try:
        url = f"https://drive.usercontent.google.com/download?id={file_id}&export=download"
        resp = requests.head(url, timeout=10, allow_redirects=True)
        lm = resp.headers.get("Last-Modified", "")
        if lm:
            dt_utc = parsedate_to_datetime(lm)
            dt_pht = dt_utc.astimezone(PHT)
            return dt_pht.strftime("%m/%d/%Y %I:%M %p")
    except Exception as e:
        logger.warning(f"Could not get Last-Modified for {file_id}: {e}")
    return ""


def parse_exp_from_batch(batch: str) -> "date | None":
    """
    Extract expiry date from a batch code.
    Format: first 6 digits = mfg date (YYMMDD), next 6 digits = expiry date (YYMMDD).
    Letters/suffixes are ignored — only digits are used.
    Returns a date object or None if not parseable.
    """
    if not batch:
        return None
    digits = ''.join(c for c in batch if c.isdigit())
    if len(digits) < 12:
        return None
    try:
        return datetime.strptime(digits[6:12], "%y%m%d").date()
    except ValueError:
        return None


def parse_exp_date(exp_raw, batch: str = "") -> "date | None":
    """
    Resolve an expiry date from either:
    1. The raw exp_date cell value (datetime / date / MM/DD/YY string)
    2. The batch code (digits 7-12 in YYMMDD format) — primary source
    Batch-derived date takes priority when available.
    """
    # Primary: derive from batch code
    batch_exp = parse_exp_from_batch(batch)
    if batch_exp is not None:
        return batch_exp
    # Fallback: use the Expiration Date column if populated
    if exp_raw is None:
        return None
    if isinstance(exp_raw, datetime):
        return exp_raw.date()
    if isinstance(exp_raw, date):
        return exp_raw
    if isinstance(exp_raw, str):
        for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(exp_raw.strip(), fmt).date()
            except ValueError:
                continue
    return None


def download_gdrive_file(file_id: str, retries: int = 3, is_sheets: bool = False) -> bytes:
    if is_sheets:
        url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    else:
        url = f"https://drive.google.com/uc?export=download&id={file_id}"
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            if b"confirm=" in resp.content[:1000]:
                match = re.search(rb'confirm=([0-9A-Za-z_\-]+)', resp.content)
                if match:
                    token = match.group(1).decode()
                    resp = requests.get(f"{url}&confirm={token}", timeout=30)
                    resp.raise_for_status()
            return resp.content
        except Exception as e:
            logger.warning(f"Download attempt {attempt}/{retries} failed for {file_id}: {e}")
            if attempt < retries:
                import time; time.sleep(2 ** attempt)
    raise RuntimeError(f"Failed to download file {file_id} after {retries} attempts")


# ──────────────────────────────────────────────────────────────────────────────
# Data loaders
# ──────────────────────────────────────────────────────────────────────────────
def load_inventory():
    logger.info("Refreshing inventory data from Google Drive...")
    data = download_gdrive_file(INVENTORY_FILE_ID)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    header = [str(h).strip() if h else "" for h in rows[0]]

    def col(name):
        try: return header.index(name)
        except ValueError: return None

    item_no_col  = col("Item No.")
    desc_col     = col("Item Description")
    stock_col    = col("Quantity")
    whs_code_col = col("Whse")
    whs_name_col = col("Whse Name")
    batch_col    = col("Batch")
    exp_col      = col("Expiration Date")
    status_col   = col("Status")

    excluded = 0
    records = []
    component_records = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        whs_name = str(row[whs_name_col]).strip() if whs_name_col is not None and row[whs_name_col] else ""
        try:
            qty = float(row[stock_col]) if stock_col is not None and row[stock_col] is not None else 0.0
        except:
            qty = 0.0
        exp_raw = row[exp_col] if exp_col is not None else None
        batch_str = str(row[batch_col]).strip() if batch_col is not None and row[batch_col] else ""
        exp_date = parse_exp_date(exp_raw, batch_str)
        status = str(row[status_col]).strip() if status_col is not None and row[status_col] else ""
        rec = {
            "item_no":   str(row[item_no_col]).strip() if item_no_col is not None and row[item_no_col] else "",
            "desc":      str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else "",
            "in_stock":  qty,
            "whs_name":  whs_name,
            "whs_code":  str(row[whs_code_col]).strip() if whs_code_col is not None and row[whs_code_col] else "",
            "batch":     batch_str,
            "exp_date":  exp_date,
            "status":    status,
        }
        if whs_name == "Component Warehouse":
            component_records.append(rec)
        elif whs_name in ON_HOLD_WAREHOUSES:
            excluded += 1
        else:
            records.append(rec)

    logger.info(f"Excluded {excluded} rows from on-hold warehouses (excl. Component Warehouse)")
    logger.info(f"Loaded {len(records)} records | {len(set(r['item_no'] for r in records))} items | {len(set(r['whs_name'] for r in records if r['whs_name']))} warehouses")
    logger.info(f"Component Warehouse: {len(component_records)} records | {len(set(r['item_no'] for r in component_records))} items")
    # Store component records in the DataStore directly (caller must hold the lock)
    store.component_inventory = component_records
    return records


def load_ar():
    logger.info("Refreshing AR data from Google Drive...")
    data = download_gdrive_file(AR_FILE_ID, is_sheets=True)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    # Build CardCode -> SEGMENT mapping from BP_ByCustomer_Segment sheet
    segment_map = {}
    try:
        ws_seg = wb["BP_ByCustomer_Segment"]
        seg_rows = list(ws_seg.iter_rows(min_row=1, values_only=True))
        seg_header = [str(h).strip() if h else "" for h in seg_rows[0]]
        seg_code_col = seg_header.index("CardCode") if "CardCode" in seg_header else None
        seg_col = seg_header.index("SEGMENT") if "SEGMENT" in seg_header else None
        if seg_code_col is not None and seg_col is not None:
            for row in seg_rows[1:]:
                if row and row[seg_code_col]:
                    segment_map[str(row[seg_code_col]).strip()] = str(row[seg_col]).strip() if row[seg_col] else ""
        logger.info(f"AR segment map loaded: {len(segment_map)} entries")
    except Exception as e:
        logger.warning(f"Could not load BP_ByCustomer_Segment: {e}")

    ws = wb["Merge1"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(h).strip() if h else "" for h in rows[0]]

    def col(name):
        try: return header.index(name)
        except ValueError: return None

    card_code_col  = col("CardCode")
    card_name_col  = col("CardName")
    agent_col      = col("Agent")
    si_date_col    = col("SI Date")
    due_date_col   = col("Due Date")
    si_num_col     = col("SI Number")
    dr_num_col     = col("DR Number")
    bir_si_col     = col("BIR SI")
    terms_col      = col("Terms")
    balance_col    = col("Balance")
    days_due_col   = col("Days Due")

    records = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        try:
            balance = float(row[balance_col]) if balance_col is not None and row[balance_col] is not None else 0.0
        except:
            balance = 0.0
        if balance <= 0:
            continue
        try:
            days_due = int(row[days_due_col]) if days_due_col is not None and row[days_due_col] is not None else 0
        except:
            days_due = 0
        card_code = str(row[card_code_col]).strip() if card_code_col is not None and row[card_code_col] else ""
        records.append({
            "card_code": card_code,
            "card_name": str(row[card_name_col]).strip() if card_name_col is not None and row[card_name_col] else "",
            "agent":     str(row[agent_col]).strip() if agent_col is not None and row[agent_col] else "",
            "area":      segment_map.get(card_code, ""),
            "si_date":   row[si_date_col] if si_date_col is not None else None,
            "due_date":  row[due_date_col] if due_date_col is not None else None,
            "si_number": str(row[si_num_col]).strip() if si_num_col is not None and row[si_num_col] else "",
            "dr_number": str(row[dr_num_col]).strip() if dr_num_col is not None and row[dr_num_col] else "",
            "bir_si":    str(row[bir_si_col]).strip() if bir_si_col is not None and row[bir_si_col] else "",
            "terms":     str(row[terms_col]).strip() if terms_col is not None and row[terms_col] else "",
            "balance":   balance,
            "days_due":  days_due,
        })

    total = sum(r["balance"] for r in records)
    clients = len(set(r["card_name"] for r in records))
    logger.info(f"AR data loaded: {len(records)} SI rows | {clients} clients | Total: ₱{total:,.2f}")
    return records


def load_ap():
    logger.info("Refreshing AP (Unreleased Payments) data from Google Drive...")
    data = download_gdrive_file(AP_FILE_ID)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Find header row (contains "DocNum")
    header_idx = None
    for i, row in enumerate(rows):
        if row and "DocNum" in [str(c) for c in row if c]:
            header_idx = i
            break
    if header_idx is None:
        logger.error("AP: Could not find header row")
        return []

    header = [str(h).strip() if h else "" for h in rows[header_idx]]

    def col(name):
        try: return header.index(name)
        except ValueError: return None

    doc_num_col    = col("DocNum")
    doc_date_col   = col("DocDate")
    due_date_col   = col("DocDueDate")
    released_col   = col("Released Date")
    card_code_col  = col("CardCode")
    card_name_col  = col("CardName")
    doc_total_col  = col("DocTotal")
    comments_col   = col("Comments")
    check_num_col  = col("CheckNum")

    records = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        released = row[released_col] if released_col is not None else None
        if released is not None and str(released).strip() not in ("", "None"):
            continue  # Skip released payments
        try:
            doc_total = float(row[doc_total_col]) if doc_total_col is not None and row[doc_total_col] is not None else 0.0
        except:
            doc_total = 0.0
        records.append({
            "doc_num":   str(row[doc_num_col]).strip() if doc_num_col is not None and row[doc_num_col] else "",
            "doc_date":  row[doc_date_col] if doc_date_col is not None else None,
            "due_date":  row[due_date_col] if due_date_col is not None else None,
            "card_code": str(row[card_code_col]).strip() if card_code_col is not None and row[card_code_col] else "",
            "card_name": str(row[card_name_col]).strip() if card_name_col is not None and row[card_name_col] else "",
            "doc_total": doc_total,
            "comments":  str(row[comments_col]).strip() if comments_col is not None and row[comments_col] else "",
            "check_num": str(row[check_num_col]).strip() if check_num_col is not None and row[check_num_col] else "",
        })

    total = sum(r["doc_total"] for r in records)
    vendors = len(set(r["card_name"] for r in records))
    logger.info(f"AP data loaded (Unreleased): {vendors} vendors | {len(records)} transactions | Total: ₱{total:,.2f}")
    return records


def refresh_all_data():
    """Reload all three data sources and update source timestamps."""
    try:
        with store._lock:
            store.inventory_source_ts = get_drive_file_modified_ts(INVENTORY_FILE_ID)
            store.inventory = load_inventory()

            store.ar_source_ts = get_drive_file_modified_ts(AR_FILE_ID, is_sheets=True)
            store.ar_rows = load_ar()

            store.ap_source_ts = get_drive_file_modified_ts(AP_FILE_ID)
            store.ap_rows = load_ap()

            store.last_refresh = datetime.now(PHT)
    except Exception as e:
        logger.error(f"Error loading data: {e}")


def refresh_ar_only():
    """Reload only AR data."""
    try:
        with store._lock:
            store.ar_source_ts = get_drive_file_modified_ts(AR_FILE_ID, is_sheets=True)
            store.ar_rows = load_ar()
            store.last_refresh = datetime.now(PHT)
    except Exception as e:
        logger.error(f"Error loading AR data: {e}")


def refresh_ap_only():
    """Reload only AP data."""
    try:
        with store._lock:
            store.ap_source_ts = get_drive_file_modified_ts(AP_FILE_ID)
            store.ap_rows = load_ap()
            store.last_refresh = datetime.now(PHT)
    except Exception as e:
        logger.error(f"Error loading AP data: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
def fmt_date(d) -> str:
    if d is None:
        return "—"
    if isinstance(d, datetime):
        return d.strftime("%m/%d/%Y")
    if isinstance(d, date):
        return d.strftime("%m/%d/%Y")
    return str(d)


def fmt_peso(amount: float) -> str:
    return f"₱{amount:,.2f}"


def inv_source_footer() -> str:
    """Footer line for inventory commands."""
    ts = store.inventory_source_ts
    if ts:
        return f"🕐 _Source data as of {ts} PHT_"
    elif store.last_refresh:
        return f"🕐 _Data as of {store.last_refresh.strftime('%m/%d/%Y %I:%M %p')} PHT_"
    return ""


def ar_source_footer() -> str:
    """Footer line for AR commands."""
    ts = store.ar_source_ts
    if ts:
        return f"🕐 _Source data as of {ts} PHT_"
    elif store.last_refresh:
        return f"🕐 _Data as of {store.last_refresh.strftime('%m/%d/%Y %I:%M %p')} PHT_"
    return ""


def ap_source_footer() -> str:
    """Footer line for AP commands."""
    ts = store.ap_source_ts
    if ts:
        return f"🕐 _Source data as of {ts} PHT_"
    elif store.last_refresh:
        return f"🕐 _Data as of {store.last_refresh.strftime('%m/%d/%Y %I:%M %p')} PHT_"
    return ""


def group_inventory_by_item(records: list[dict]) -> dict:
    """Group inventory records by item_no, summing stock per warehouse."""
    by_item: dict[str, dict] = {}
    for r in records:
        k = r["item_no"]
        if k not in by_item:
            by_item[k] = {
                "item_no": k,
                "desc": r["desc"],
                "total": 0.0,
                "warehouses": {},
                "batches": [],
            }
        by_item[k]["total"] += r["in_stock"]
        whs = r["whs_name"]
        if whs:
            by_item[k]["warehouses"][whs] = by_item[k]["warehouses"].get(whs, 0.0) + r["in_stock"]
        if r.get("exp_date"):
            by_item[k]["batches"].append({"exp_date": r["exp_date"], "qty": r["in_stock"]})
    return by_item


def search_inventory(query: str) -> list[dict]:
    """Search inventory by keyword in description or item_no."""
    q = query.lower().strip()
    records = store.inventory
    exact = [r for r in records if r["item_no"].lower() == q]
    if exact:
        return exact
    matches = [r for r in records if q in r["desc"].lower() or q in r["item_no"].lower()]
    return matches


def get_category_records(category: str) -> list[dict]:
    """Return inventory records matching a product category."""
    cat_upper = category.upper()
    keywords = CATEGORY_KEYWORDS.get(cat_upper, [cat_upper.lower()])
    matches = []
    for r in store.inventory:
        desc_lower = r["desc"].lower()
        if any(kw in desc_lower for kw in keywords):
            matches.append(r)
    return matches


def get_ap_days_overdue(r: dict) -> int:
    """Calculate how many days an AP record is overdue (positive = overdue)."""
    today = datetime.now(PHT).date()
    d = r.get("due_date")
    if d is None:
        return 0
    if isinstance(d, datetime):
        d = d.date()
    if isinstance(d, date):
        return (today - d).days
    return 0


def correct_command_typo(text: str) -> str | None:
    """Return the corrected command if the text looks like a typo of a known command."""
    known = [
        "/search", "/check", "/refresh", "/summary", "/category", "/expiring",
        "/low", "/warehouse", "/slowmoving", "/top", "/components", "/components_expiring",
        "/ar", "/client", "/aging", "/overdue", "/area", "/agent",
        "/arsearch", "/arsummary", "/arrefresh",
        "/ap", "/vendor", "/ap_summary", "/apsummary", "/apaging",
        "/apoverdue", "/aptop", "/due_today", "/aprefresh",
        "/start", "/help", "/myid",
        "/register", "/unregister", "/listusers", "/seen",
        "/accessmode", "/whitelistgroup", "/unallowgroup",
        "/allowedgroups", "/seengroups",
    ]
    text = text.strip().lower()
    if not text.startswith("/"):
        return None
    cmd = text.split()[0]
    if cmd in known:
        return None
    matches = difflib.get_close_matches(cmd, known, n=1, cutoff=0.7)
    if matches:
        rest = text[len(cmd):]
        return matches[0] + rest
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Command handlers — General
# ──────────────────────────────────────────────────────────────────────────────
HELP_TEXT = (
    "👋 *Belcris Inventory Bot*\n\n"
    "*📦 Inventory:*\n"
    "• `/refresh` — Reload all data from Google Drive\n"
    "• `/summary` — Inventory snapshot (all warehouses)\n"
    "• `/search <keyword>` — Search items by name\n"
    "• `/check <item_code>` — Look up item by SAP code\n"
    "• `/category [name]` — Browse by category (HAM, BACON, TOCINO…)\n"
    "• `/expiring [days]` — Items expiring within N days (default 30)\n"
    "• `/lowstock [keyword] [threshold]` — Items below threshold (default 500 kg)\n"
    "• `/warehouse <name or code>` — Stock in a specific warehouse\n"
    "• `/top` — Top 20 items by quantity\n"
    "• `/slowmoving [days]` — Items with stock but no sales (default 30 days)\n"
    "• `/components [keyword]` — Component Warehouse stock with expiry dates (PRD use)\n"
    "• `/components_expiring [days]` — Components expiring within N days (default 30)\n\n"
    "*📋 Accounts Receivable:*\n"
    "• `/arsummary` — AR executive summary + top clients & areas\n"
    "• `/client <name or code>` — Full AR detail for a client\n"
    "  _Tip: use client code (e.g. CMNL00013) for direct lookup_\n"
    "• `/aging` — AR aging breakdown by bucket\n"
    "• `/overdue [days]` — Top overdue clients (default 1+ days)\n"
    "• `/area <name>` — Receivables grouped by area\n"
    "• `/agent <name>` — All clients under a sales agent\n"
    "• `/arsearch <keyword>` — Search clients by name or code\n"
    "• `/arrefresh` — Refresh AR data only\n\n"
    "*💳 Accounts Payable:*\n"
    "• `/apsummary` — AP summary with aging breakdown\n"
    "• `/ap <vendor>` — Vendor AP balance & transaction detail\n"
    "• `/apaging` — AP aging by bucket\n"
    "• `/apoverdue` — Vendors overdue 61+ days\n"
    "• `/aptop [n]` — Top N vendors by outstanding amount\n"
    "• `/due_today` — AP payments due this week\n"
    "• `/aprefresh` — Refresh AP data only\n\n"
    "*📈 Sales Performance:*\n"
    "• `/sales [area] [month]` — Sales by area (optional: filter by area & month)\n"
    "  _e.g. `/sales cebu july`, `/sales manila 7`, `/sales july`_\n"
    "• `/salesagent [name]` — Revenue by sales agent\n"
    "• `/salestarget [month]` — Segment performance vs target\n"
    "• `/salesproduct [keyword] [cat:CATEGORY]` — Top products by revenue\n"
    "• `/salesmonth` — Monthly revenue trend\n\n"
    "*💡 Tips:*\n"
    "• Prefix `//` to any message to silently ignore it\n"
    "• Commands are case-insensitive\n"
    "• `/myid` — get your Telegram User ID (share with admin)\n"
)

ADMIN_HELP_ENTRIES = (
    "*🔐 Admin Commands* (for admins only):\n"
    "• `/register <id1>,<id2>,...` — Batch register users by Telegram ID\n"
    "• `/unregister <id>` — Remove a registered user\n"
    "• `/listusers` — List all registered users\n"
    "• `/testreports` — Manually trigger automated reports (Admin only)\n"
    "• `/seen` — Show users who used the bot but are not registered\n"
    "• `/accessmode [off|soft|hard]` — View/set access enforcement mode\n"
    "• `/whitelistgroup [chat_id]` — Fully open the current (or given) group chat for all members\n"
    "• `/unallowgroup [chat_id]` — Remove a group from the open list\n"
    "• `/allowedgroups` — List open groups\n"
    "• `/seengroups` — List all groups the bot has been used in\n"
)


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/start — plain help for everyone; admin section appended for admins only."""
    user = update.effective_user
    log_user(user.id if user else None, user.username if user else None, user.full_name if user else None)
    text = HELP_TEXT
    if _is_admin(user.id if user else None):
        text += "\n" + ADMIN_HELP_ENTRIES
    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)


async def cmd_myid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Return the user's Telegram ID and the current chat ID. Also passively logs the user."""
    user = update.effective_user
    chat = update.effective_chat
    log_user(user.id if user else None, user.username if user else None, user.full_name if user else None)

    user_id = user.id if user else "unknown"
    user_name = user.full_name if user else "unknown"
    username = f"@{user.username}" if user and user.username else "(no username)"
    chat_id = chat.id if chat else "unknown"
    chat_type = chat.type if chat else "unknown"
    chat_title = chat.title if chat and chat.title else "(private chat)"

    # Register self if this user is an admin
    admin_status = "✅ Admin — full access" if _is_admin(user.id) else ""

    lines = [
        "🪪 *Your Telegram Info*",
        "",
        f"👤 Name: {user_name}",
        f"🔖 Username: {username}",
        f"🆔 Your User ID: `{user_id}`",
        "",
        f"💬 Chat Type: {chat_type}",
        f"📛 Chat Title: {chat_title}",
        f"🆔 Chat ID: `{chat_id}`",
    ]
    if admin_status:
        lines.append("")
        lines.append(admin_status)
    lines.append("")
    lines.append("_Share your User ID with the admin for bot access registration._")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/help — shows admin commands only to users in ADMIN_IDS."""
    user = update.effective_user
    log_user(user.id if user else None, user.username if user else None, user.full_name if user else None)
    text = HELP_TEXT
    if _is_admin(user.id if user else None):
        text += "\n" + ADMIN_HELP_ENTRIES
    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)


# ──────────────────────────────────────────────────────────────────────────────
# Command handlers — Admin / Access Control
# Admin-only: requires the caller's Telegram user ID to be in the ADMIN_IDS
# Railway env var. Admin commands themselves are never access-gated.
# ──────────────────────────────────────────────────────────────────────────────
ADMIN_HELP = (
    "🔐 *Admin Commands*\n\n"
    "• `/register <id> [note]` — Register a user by Telegram ID\n"
    "  _or_ `/register <id1>,<id2>,...` — Batch register multiple IDs\n"
    "• `/unregister <id>` — Remove a registered user\n"
    "• `/listusers` — List all registered users\n"
    "• `/seen` — List users who have interacted with the bot (not yet registered)\n"
    "• `/accessmode [off|soft|hard]` — Set the access enforcement mode (admin only)\n"
    "• `/summarize [chat_id]` — Summarize recent chat and extract tasks (Admin only)\n"
    "• `/myid` — Show your own Telegram User ID (needed to register yourself)\n\n"
    "_Admin IDs are set by the `ADMIN_IDS` Railway variable._"
)


def _parse_id_list(args: list[str]) -> list[int]:
    """Parse comma/space-separated Telegram IDs from args, e.g. /register 123,456 789."""
    ids: list[int] = []
    for arg in args:
        for part in arg.replace(",", " ").split():
            if part.isdigit():
                ids.append(int(part))
    return ids


async def cmd_register(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/register <id1>[,<id2>,...] [note] — register users (batch supported).
    Optional trailing note applies to all IDs in this call.
    Also auto-registers any of the given IDs that already appear in the seen log."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    args = context.args or []
    if not args:
        await update.message.reply_text(
            "Usage: `/register <id> [note]`\n"
            "Batch: `/register <id1>,<id2>,<id3> [note]`\n"
            "_Get IDs via `/myid`._",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    # Numeric tokens (with optional commas) are IDs; the first non-numeric token
    # starts the optional trailing note that applies to all IDs in the call.
    note = None
    raw_ids: list[int] = []
    for a in args:
        if a.replace(",", "").isdigit():
            raw_ids.extend(int(p) for p in a.split(",") if p.isdigit())
        else:
            note = " ".join(args[args.index(a):])
            break

    if not raw_ids:
        await update.message.reply_text("❌ No valid Telegram IDs found. Use `/myid` to get an ID.", parse_mode=ParseMode.MARKDOWN)
        return

    seen_map = {r["telegram_user_id"]: r for r in list_seen_users(limit=10000)}
    added = []
    already = []
    for uid in dict.fromkeys(raw_ids):
        if register_user(uid, admin.id, note):
            added.append(uid)
        else:
            already.append(uid)

    lines = ["🔐 *Registration Complete*"]
    for uid in added:
        known = seen_map.get(uid)
        name = known["full_name"] if known and known["full_name"] else "(unknown name)"
        uname = known["username"] if known and known["username"] else ""
        extra = f" {name} {uname}".strip() if (known and (known["full_name"] or known["username"])) else ""
        lines.append(f"✅ Registered `{uid}`{(' — ' + extra) if extra else ''}")
    for uid in already:
        lines.append(f"⚠️ `{uid}` was already registered")
    if note:
        lines.append(f"_Note: {note}_")
    lines.append(f"\nTotal registered users: {len(list_registered_users())}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_unregister(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/unregister <id> — remove a user from the registered list."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    args = context.args or []
    if not args or not args[0].isdigit():
        await update.message.reply_text("Usage: `/unregister <user_id>`", parse_mode=ParseMode.MARKDOWN)
        return
    uid = int(args[0])
    if unregister_user(uid):
        await update.message.reply_text(f"✅ User `{uid}` has been removed from the registered list.", parse_mode=ParseMode.MARKDOWN)
    else:
        await update.message.reply_text(f"❌ User `{uid}` was not registered.", parse_mode=ParseMode.MARKDOWN)


async def cmd_testreports(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/testreports — Manually trigger automated reports for testing (Admin only)."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    
    await update.message.reply_text("🧪 Triggering automated reports for all whitelisted groups...")
    asyncio.run_coroutine_threadsafe(send_automated_report("daily_inventory"), _loop)
    asyncio.run_coroutine_threadsafe(send_automated_report("weekly_collections"), _loop)
    asyncio.run_coroutine_threadsafe(send_automated_report("weekly_slowmoving"), _loop)


async def cmd_summarize(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/summarize [chat_id] — Summarize recent conversation and extract tasks (Admin only)."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return

    # If used in a group, summarize that group. If in private, requires chat_id.
    chat = update.effective_chat
    target_chat_id = chat.id
    if chat.type == "private":
        args = context.args or []
        if not args or not args[0].lstrip("-").isdigit():
            await update.message.reply_text(
                "Usage: `/summarize <chat_id>` (get IDs from `/seengroups`)",
                parse_mode=ParseMode.MARKDOWN
            )
            return
        target_chat_id = int(args[0])

    if not is_whitelisted_group(target_chat_id):
        await update.message.reply_text("❌ This chat is not whitelisted for summarization.")
        return

    # Fetch last 100 messages
    try:
        with _db_lock:
            conn = _access_db()
            rows = conn.execute(
                "SELECT full_name, message_text, timestamp FROM group_messages WHERE telegram_chat_id = ? ORDER BY id ASC LIMIT 100",
                (target_chat_id,)
            ).fetchall()
            conn.close()
    except Exception as e:
        await update.message.reply_text(f"❌ Database error: {e}")
        return

    if not rows:
        await update.message.reply_text("📭 No recent messages found in this chat to summarize.")
        return

    convo = "\n".join([f"[{r['timestamp']}] {r['full_name']}: {r['message_text']}" for r in rows])
    
    msg = await update.message.reply_text("🧠 Analyzing conversation and extracting tasks...")

    try:
        # Use OpenAI to summarize
        api_key = os.environ.get("OPENAI_API_KEY")
        base_url = os.environ.get("OPENAI_API_BASE")
        if not api_key:
            await msg.edit_text("❌ OpenAI API key not configured on Railway.")
            return

        response = requests.post(
            f"{base_url}/chat/completions",
            headers={"Authorization": f"Bearer {api_key}"},
            json={
                "model": "gpt-4o",
                "messages": [
                    {"role": "system", "content": "You are a helpful assistant for Belcris. Summarize the following Telegram conversation and extract any assigned tasks or action items. Format with clear headings and bullet points. Keep it professional."},
                    {"role": "user", "content": f"Please summarize this conversation and list tasks:\n\n{convo}"}
                ]
            },
            timeout=30
        )
        data = response.json()
        summary = data["choices"][0]["message"]["content"]
        
        await msg.edit_text(f"📝 *Conversation Summary & Tasks*\n\n{summary}", parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        logger.error(f"Summarization error: {e}")
        await msg.edit_text(f"❌ Failed to summarize: {e}")


async def cmd_listusers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/listusers — list all registered users with their Telegram IDs."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    rows = list_registered_users()
    if not rows:
        await update.message.reply_text("📋 No registered users yet.\nUse `/register <id>` or batch `/register <id1>,<id2>`.")
        return
    lines = [f"📋 *Registered Users* ({len(rows)})"]
    for r in rows:
        name = r["full_name"] or r["username"] or "(no name logged yet)"
        name_safe = name.replace("*", "").replace("_", "")
        last = r["last_seen"] or "never"
        note = f" ({r['note']})" if r["note"] else ""
        lines.append(f"• `{r['telegram_user_id']}` — {name_safe}{note}\n  _Last seen: {last} PHT_")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_seen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/seen — list users who have interacted with the bot but are not yet registered.
    Useful for batch registration from the collected IDs."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    registered = {r["telegram_user_id"] for r in list_registered_users()}
    rows = [r for r in list_seen_users(limit=10000) if r["telegram_user_id"] not in registered]
    if not rows:
        await update.message.reply_text("✅ All known users are already registered (or no one has interacted yet).")
        return
    lines = [f"👀 *Seen, Not Registered* ({len(rows)})\n_Reply `/register` with these IDs to register them in batch._", ""]
    chunk = "\n".join(lines)
    for r in rows:
        name = r["full_name"] or r["username"] or "(no name)"
        name_safe = name.replace("*", "").replace("_", "")
        line = f"• `{r['telegram_user_id']}` — {name_safe} _(seen {r['last_seen']} PHT, {r['message_count']} msgs)_\n"
        if len(chunk) + len(line) > 3800:
            await update.message.reply_text(chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)
            chunk = line
        else:
            chunk += line
    await update.message.reply_text(chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)


async def cmd_accessmode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/accessmode [off|soft|hard] — show or set access enforcement mode (admin only).
    Setting it here takes effect immediately for this process. To make it survive
    Railway restarts, also set the `ACCESS_MODE` Railway variable."""
    global ACCESS_MODE  # noqa: PLW0603
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    args = context.args or []
    if args and args[0].lower() in ("off", "soft", "hard"):
        ACCESS_MODE = args[0].lower()
    lines = [
        "🔒 *Access Control Status*",
        "",
        f"Current mode: *{ACCESS_MODE.upper()}*",
        "",
        "• `off` — open access, IDs are just logged",
        "• `soft` — warn unregistered users, still allow data",
        "• `hard` — block unregistered users from data commands",
        "",
    ]
    if args and args[0].lower() in ("off", "soft", "hard"):
        lines.append(f"✅ Mode set to *{ACCESS_MODE.upper()}* for this session.")
        lines.append("_Note: Railway must set `ACCESS_MODE` permanently (redeploy persists it across restarts)._")
    else:
        lines.append("Usage: `/accessmode off|soft|hard`")
    lines.append("")
    lines.append(f"Admin IDs configured: {sorted(ADMIN_IDS) if ADMIN_IDS else '_none (set the `ADMIN_IDS` Railway variable)_'}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_whitelistgroup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/whitelistgroup [chat_id] — Fully open a group chat for all members (no action needed from them).
    With no arg, whitelists the current chat. With an arg, whitelists the given Chat ID
    (useful when run in a private chat — the target group Chat ID is a negative number)."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    chat = update.effective_chat
    args = context.args or []
    if args:
        raw = args[0].replace(",", "").strip()
        if not (raw.lstrip("-").isdigit()):
            await update.message.reply_text("❌ Chat ID must be a number (groups are negative, e.g. `-1001234567890`).")
            return
        target_id = int(raw)
        if target_id >= 0:
            await update.message.reply_text("⚠️ Positive IDs are users, not groups. Use a negative group Chat ID (from `/myid` in that group).")
            return
    else:
        if chat is None or chat.id is None:
            await update.message.reply_text("❌ Could not determine the current chat.")
            return
        target_id = chat.id

    if allow_group(target_id, admin.id):
        title = chat.title if chat is not None else "(target group)"
        await update.message.reply_text(
            f"✅ Group *`{target_id}`* is now fully open for all members.\n"
            f"No registration or action needed from anyone in it (including your boss).\n"
            f"_(Chat: {title.replace('*', '').replace('_', '') if title else 'added by ID'})_",
            parse_mode=ParseMode.MARKDOWN,
        )
    else:
        await update.message.reply_text(f"⚠️ Group *`{target_id}`* is already on the open list.", parse_mode=ParseMode.MARKDOWN)


async def cmd_unallowgroup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/unallowgroup <chat_id> — Remove a group from the open list."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    args = context.args or []
    if not args:
        await update.message.reply_text("Usage: `/unallowgroup <group_chat_id>`", parse_mode=ParseMode.MARKDOWN)
        return
    raw = args[0].replace(",", "").strip()
    if not raw.lstrip("-").isdigit():
        await update.message.reply_text("❌ Chat ID must be a number.")
        return
    target_id = int(raw)
    if unallow_group(target_id):
        await update.message.reply_text(f"✅ Group *`{target_id}`* removed from the open list.", parse_mode=ParseMode.MARKDOWN)
    else:
        await update.message.reply_text(f"❌ Group *`{target_id}`* was not on the open list.", parse_mode=ParseMode.MARKDOWN)


async def cmd_allowedgroups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/allowedgroups — List groups that are fully open for all members."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    rows = list_allowed_groups()
    if not rows:
        await update.message.reply_text(
            "📋 No groups are currently on the open list.\n"
            "Send `/whitelistgroup` in the group chat to open it for all members.\n"
            "Or set `ALLOWED_GROUP_CHAT_IDS` in Railway variables."
        )
        return
    lines = [f"🟢 *Open Groups* ({len(rows)})\n_All members can use the bot freely in these chats._"]
    for r in rows:
        title = r["chat_title"] or "(no title logged)"
        title_safe = title.replace("*", "").replace("_", "")
        note = f" ({r['note']})" if r["note"] else ""
        lines.append(f"• *`{r['telegram_chat_id']}`* — {title_safe}{note}\n  _Added: {r['allowed_at']} PHT_")
    if ALLOWED_GROUP_CHAT_IDS:
        env_ids = [f"`{i}`" for i in sorted(ALLOWED_GROUP_CHAT_IDS)]
        lines.append("")
        lines.append(f"Also open via Railway `ALLOWED_GROUP_CHAT_IDS`: {', '.join(env_ids)}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_seengroups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/seengroups — List all group chats the bot has been used in (not yet open)."""
    admin = update.effective_user
    if not _is_admin(admin.id):
        await update.message.reply_text("🚫 This command is admin-only.")
        return
    allowed = {r["telegram_chat_id"] for r in list_allowed_groups()} | ALLOWED_GROUP_CHAT_IDS
    rows = [r for r in list_seen_groups(limit=10000) if r["telegram_chat_id"] not in allowed]
    if not rows:
        await update.message.reply_text("✅ All known groups are already on the open list (or no groups seen yet).")
        return
    lines = [f"👀 *Groups Seen, Not Open* ({len(rows)})\n_Send `/whitelistgroup` in any of these to open it for all members._", ""]
    chunk = "\n".join(lines)
    for r in rows:
        title = (r["chat_title"] or f"{r['chat_type']} chat").replace("*", "").replace("_", "")
        line = f"• *`{r['telegram_chat_id']}`* — {title} _(seen {r['last_seen']} PHT, {r['interaction_count']} interactions)_\n"
        if len(chunk) + len(line) > 3800:
            await update.message.reply_text(chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)
            chunk = line
        else:
            chunk += line
    await update.message.reply_text(chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)


async def cmd_refresh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    msg = await update.message.reply_text("🔄 Refreshing data from Google Drive...")
    try:
        refresh_all_data()
        now_pht = datetime.now(PHT)
        ts_date = store.last_refresh.strftime("%m/%d/%Y") if store.last_refresh else "—"
        ts_full = store.last_refresh.strftime("%m/%d/%Y %I:%M %p") if store.last_refresh else "—"

        inv = store.inventory
        total_records = len(inv)
        unique_items  = len(set(r["item_no"] for r in inv))
        warehouses    = len(set(r["whs_name"] for r in inv if r["whs_name"]))
        total_qty     = int(sum(r["in_stock"] for r in inv))

        # Expiry alerts — exp_date is always a date object (parsed from batch code)
        expired_items = set()
        exp_30_items  = set()
        exp_90_items  = set()
        today_pht = now_pht.date()
        for r in inv:
            exp_date = r.get("exp_date")
            if exp_date is None:
                continue
            days_left = (exp_date - today_pht).days
            if r["in_stock"] > 0 and days_left < 0:
                expired_items.add(r["item_no"])
            elif days_left <= 30:
                exp_30_items.add(r["item_no"])
            elif days_left <= 90:
                exp_90_items.add(r["item_no"])

        src_ts = store.inventory_source_ts or ts_full

        lines = [
            f"✅ Data refreshed successfully!",
            f"🏪 *Inventory Summary*",
            f"_As of {src_ts}_",
            f"",
            f"Total Records: {total_records:,}",
            f"Unique Items: {unique_items:,}",
            f"Warehouses: {warehouses}",
            f"Total Quantity: {total_qty:,}",
            f"",
            f"*Expiry Alerts:*",
            f"🔴 Expired (still in stock): {len(expired_items)} items",
            f"🟠 Expiring ≤30 days: {len(exp_30_items)} items",
            f"🟡 Expiring ≤90 days: {len(exp_90_items)} items",
            f"",
            f"🕐 _Source data as of {src_ts} PHT_",
        ]
        await msg.edit_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        await msg.edit_text(f"❌ Refresh failed: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Command handlers — Inventory
# ──────────────────────────────────────────────────────────────────────────────
async def cmd_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inventory snapshot: total stock, category breakdown, expiry alerts."""
    if not await _access_gate(update, context):
        return
    if not store.inventory:
        await update.message.reply_text("⚠️ Inventory data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    now_pht = datetime.now(PHT)
    inv = store.inventory
    total_records = len(inv)
    unique_items  = len(set(r["item_no"] for r in inv))
    warehouses    = len(set(r["whs_name"] for r in inv if r["whs_name"]))
    total_qty     = int(sum(r["in_stock"] for r in inv))

    # Category breakdown
    cat_totals: dict[str, int] = {}
    for cat, keywords in CATEGORY_KEYWORDS.items():
        total = sum(r["in_stock"] for r in inv if any(kw in r["desc"].lower() for kw in keywords))
        if total > 0:
            cat_totals[cat] = int(total)

    # Expiry alerts — exp_date is always a date object (parsed from batch code)
    expired_items = set()
    exp_30_items  = set()
    exp_90_items  = set()
    today_pht = now_pht.date()
    for r in inv:
        exp_date = r.get("exp_date")
        if exp_date is None:
            continue
        days_left = (exp_date - today_pht).days
        if r["in_stock"] > 0 and days_left < 0:
            expired_items.add(r["item_no"])
        elif days_left <= 30:
            exp_30_items.add(r["item_no"])
        elif days_left <= 90:
            exp_90_items.add(r["item_no"])

    lines = [
        "🏪 *Inventory Snapshot*\n",
        f"Total Records: {total_records:,}",
        f"Unique Items: {unique_items:,}",
        f"Warehouses: {warehouses}",
        f"Total Quantity: {total_qty:,}",
        "",
        "*Category Breakdown:*",
    ]
    for cat, qty in sorted(cat_totals.items(), key=lambda x: x[1], reverse=True):
        lines.append(f"• {cat}: {qty:,}")

    lines += [
        "",
        "*Expiry Alerts:*",
        f"🔴 Expired (in stock): {len(expired_items)} items",
        f"🟠 Expiring ≤30 days: {len(exp_30_items)} items",
        f"🟡 Expiring ≤90 days: {len(exp_90_items)} items",
        "",
        inv_source_footer(),
    ]
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_search (original body unchanged below)
    query = " ".join(context.args).strip() if context.args else ""
    if not query:
        await update.message.reply_text("Usage: `/search <keyword>`\nExample: `/search chicken`", parse_mode=ParseMode.MARKDOWN)
        return
    await _do_search(update, query)


async def cmd_check(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_check
    query = " ".join(context.args).strip() if context.args else ""
    if not query:
        await update.message.reply_text("Usage: `/check <item_code>`\nExample: `/check FESCHKN000`", parse_mode=ParseMode.MARKDOWN)
        return

    records = store.inventory
    exact = [r for r in records if r["item_no"].lower() == query.lower()]
    if not exact:
        if " " in query or not re.match(r'^[A-Za-z0-9\-_]+$', query):
            await update.message.reply_text(
                f"ℹ️ `{query}` doesn't look like an item code. Searching by keyword instead...",
                parse_mode=ParseMode.MARKDOWN,
            )
            await _do_search(update, query)
            return
        all_codes = list(set(r["item_no"] for r in records))
        close = difflib.get_close_matches(query.upper(), all_codes, n=3, cutoff=0.7)
        if close:
            await update.message.reply_text(
                f"❓ Item code `{query}` not found. Did you mean:\n" +
                "\n".join(f"• `{c}`" for c in close),
                parse_mode=ParseMode.MARKDOWN,
            )
            return
        await update.message.reply_text(f"❌ Item code `{query}` not found.", parse_mode=ParseMode.MARKDOWN)
        return

    await _do_item_detail(update, exact)


async def _do_search(update: Update, query: str):
    """Compact flat-list search results. Use /check for full item detail."""
    # Passive logging + soft/hard access gating for free-text searches
    user = update.effective_user
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if ACCESS_MODE == "hard":
                await _send_blocked_notice(update, str(user.id))
                return
            if ACCESS_MODE == "soft":
                await _send_blocked_notice(update, str(user.id))
                # soft: continue to results below
    if not store.inventory:
        await update.message.reply_text("⚠️ Inventory data not loaded yet. Try /refresh.")
        return

    matches = search_inventory(query)
    if not matches:
        await update.message.reply_text(f"❌ No results for '{query}'.")
        return

    by_item = group_inventory_by_item(matches)
    items = sorted(by_item.values(), key=lambda x: x["total"], reverse=True)

    ts = store.inventory_source_ts or (store.last_refresh.strftime('%m/%d/%Y %I:%M %p PHT') if store.last_refresh else '')
    footer = f"\n🕐 Source data as of {ts}" if ts else ""

    header = f'🔍 Search Results for "{query.upper()}" ({len(items)} items found)\n'
    chunk = header
    for item in items:
        qty = int(item["total"])
        desc_safe = item['desc'].replace('*', '').replace('_', '').replace('`', '').replace('[', '').replace(']', '')
        # Item code in monospace so desktop users can click-to-copy
        line = f"• `{item['item_no']}` — {desc_safe}: {qty:,}\n"
        if len(chunk) + len(line) > 3800:
            try:
                await update.message.reply_text(chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)
            except Exception as e:
                logger.error(f"_do_search chunk send failed: {e}")
                await update.message.reply_text(chunk.rstrip())
            chunk = line
        else:
            chunk += line

    chunk += f"{footer}"

    # Build inline buttons for top 10 items: show item code + compact description
    top10 = items[:10]
    buttons = []
    for item in top10:
        desc = item['desc']
        # Try to extract brand/supplier: text after last ' - ' separator
        if ' - ' in desc:
            parts = desc.rsplit(' - ', 1)
            brand = parts[1].strip()[:18]  # brand name after last dash
            core = parts[0].strip()[:18]   # product name before last dash
            short_desc = f"{core} [{brand}]"
        else:
            short_desc = desc[:28].strip()
        btn_label = f"{item['item_no']}  {short_desc}"
        buttons.append([InlineKeyboardButton(btn_label, callback_data=f"check|{item['item_no']}")])
    row = []

    keyboard = InlineKeyboardMarkup(buttons) if buttons else None
    try:
        await update.message.reply_text(chunk.rstrip(), parse_mode=ParseMode.MARKDOWN, reply_markup=keyboard)
    except Exception as e:
        logger.error(f"_do_search final chunk failed (markdown): {e}")
        try:
            await update.message.reply_text(chunk.rstrip(), reply_markup=keyboard)
        except Exception as e2:
            logger.error(f"_do_search final chunk failed (plain): {e2}")
            await update.message.reply_text(f"Found {len(items)} results for '{query}'.")


async def _do_item_detail(update: Update, records: list):
    """Rich single-item detail: batches, mfg/expiry dates, warehouse breakdown."""
    today = datetime.now(PHT).date()
    item_no = records[0]["item_no"]
    desc = records[0]["desc"]
    total_qty = sum(r["in_stock"] for r in records)

    # Warehouse breakdown
    whs_totals: dict[str, float] = {}
    for r in records:
        if r["whs_name"]:
            whs_totals[r["whs_name"]] = whs_totals.get(r["whs_name"], 0.0) + r["in_stock"]
    whs_sorted = sorted(whs_totals.items(), key=lambda x: x[1], reverse=True)

    # Batch details — group by (batch, whs_name)
    batch_rows: dict[tuple, dict] = {}
    for r in records:
        key = (r["batch"], r["whs_name"])
        if key not in batch_rows:
            batch_rows[key] = {
                "batch": r["batch"],
                "whs": r["whs_name"],
                "qty": 0.0,
                "exp_date": r["exp_date"],
            }
        batch_rows[key]["qty"] += r["in_stock"]

    # Nearest expiry
    exp_dates = [v["exp_date"] for v in batch_rows.values() if v["exp_date"] is not None]
    nearest_exp = min(exp_dates) if exp_dates else None

    def exp_tag(exp_date) -> str:
        if exp_date is None:
            return ""
        days = (exp_date - today).days
        if days < 0:
            return f"🔴 EXPIRED ({abs(days)}d ago)"
        elif days <= 30:
            return f"🟠 {days}d left"
        elif days <= 90:
            return f"🟡 {days}d left"
        else:
            return f"🟢 {days}d left"

    # Parse mfg date from batch
    def parse_mfg_from_batch(batch: str):
        digits = ''.join(c for c in batch if c.isdigit())
        if len(digits) >= 6:
            try:
                return datetime.strptime(digits[0:6], "%y%m%d").date()
            except ValueError:
                pass
        return None

    lines = [
        f"📦 *{desc}*",
        f"Item Code: `{item_no}`",
        f"Total Quantity: *{int(total_qty):,}*",
        f"Batches: {len(batch_rows)} | Warehouses: {len(whs_totals)}",
    ]
    if nearest_exp:
        days_to_exp = (nearest_exp - today).days
        lines.append(f"{exp_tag(nearest_exp).split()[0]} Nearest Expiry: {nearest_exp.strftime('%m/%d/%Y')} ({abs(days_to_exp)} days {'left' if days_to_exp >= 0 else 'ago'})")

    lines.append("")
    lines.append("*Warehouse Breakdown:*")
    for whs, qty in whs_sorted:
        lines.append(f"• {whs}: {int(qty):,}")

    BATCH_PREVIEW = 20
    batch_list = sorted(batch_rows.values(), key=lambda x: (x["exp_date"] or date.max, x["whs"]))
    total_batches = len(batch_list)
    show_all = getattr(_do_item_detail, '_show_all', False)
    display_batches = batch_list  # will be sliced below

    def build_batch_lines(batches):
        blines = []
        for b in batches:
            mfg = parse_mfg_from_batch(b["batch"])
            exp = b["exp_date"]
            mfg_str = mfg.strftime("%m/%d/%Y") if mfg else "—"
            exp_str = exp.strftime("%m/%d/%Y") if exp else "—"
            tag = exp_tag(exp)
            blines.append(f"• `{b['batch']}` | Qty: {int(b['qty']):,} | {b['whs']}")
            blines.append(f"  Mfg: {mfg_str} | Exp: {exp_str} {tag}")
        return blines

    has_more = total_batches > BATCH_PREVIEW
    preview_batches = batch_list[:BATCH_PREVIEW] if has_more else batch_list

    lines.append("")
    lines.append("*Batch Details:*")
    lines.extend(build_batch_lines(preview_batches))
    if has_more:
        lines.append(f"_...and {total_batches - BATCH_PREVIEW} more batches_")
    lines.append(f"\n{inv_source_footer()}")

    text = "\n".join(lines)
    MAX_LEN = 4000
    # Truncate if still too long (shouldn't happen with 20 batch preview)
    if len(text) > MAX_LEN:
        text = text[:MAX_LEN]

    if has_more:
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton(
            f"📋 Show All Batches ({total_batches})", callback_data=f"check_all|{item_no}"
        )]])
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=keyboard)
    else:
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)


async def cmd_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_category
    """Browse inventory by product category."""
    cat_arg = " ".join(context.args).strip().upper() if context.args else ""

    if not cat_arg:
        # List available categories with totals
        if not store.inventory:
            await update.message.reply_text("⚠️ Inventory data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
            return
        lines = ["📂 *Product Categories*\n"]
        for cat, keywords in CATEGORY_KEYWORDS.items():
            total = int(sum(r["in_stock"] for r in store.inventory if any(kw in r["desc"].lower() for kw in keywords)))
            items = len(set(r["item_no"] for r in store.inventory if any(kw in r["desc"].lower() for kw in keywords)))
            if total > 0:
                lines.append(f"• *{cat}*: {total:,} units ({items} SKUs)")
        lines.append(f"\nUsage: `/category HAM`")
        lines.append(f"\n{inv_source_footer()}")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)
        return

    if not store.inventory:
        await update.message.reply_text("⚠️ Inventory data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    matches = get_category_records(cat_arg)
    if not matches:
        cats = ", ".join(CATEGORY_KEYWORDS.keys())
        await update.message.reply_text(
            f"❌ No items found for category *{cat_arg}*.\nAvailable: {cats}",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    by_item = group_inventory_by_item(matches)
    items = sorted(by_item.values(), key=lambda x: x["total"], reverse=True)
    total_qty = int(sum(v["total"] for v in items))

    lines = [f"📂 *Category: {cat_arg}* — {len(items)} SKUs | {total_qty:,} units\n"]
    for item in items[:20]:
        if item["total"] > 0:
            lines.append(f"• *{item['item_no']}* — {item['desc'][:45]}: *{item['total']:,.0f}*")

    if len(items) > 20:
        lines.append(f"\n_Showing 20 of {len(items)} SKUs._")

    lines.append(f"\n{inv_source_footer()}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_expiring(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_expiring
    """Items expiring in N days (default 30)."""
    days = 30
    if context.args:
        try:
            days = int(context.args[0])
        except ValueError:
            await update.message.reply_text("Usage: `/expiring [days]`\nExample: `/expiring 60`", parse_mode=ParseMode.MARKDOWN)
            return

    if not store.inventory:
        await update.message.reply_text("⚠️ Inventory data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    now_pht = datetime.now(PHT)
    today = now_pht.date()
    cutoff = today + timedelta(days=days)

    expiring = []
    for r in store.inventory:
        if r["in_stock"] <= 0:
            continue
        exp_date = r.get("exp_date")  # always a date object from parse_exp_date
        if exp_date is None:
            continue
        days_left = (exp_date - today).days
        if days_left <= days:
            expiring.append({**r, "days_left": days_left, "exp_date_obj": exp_date})

    if not expiring:
        await update.message.reply_text(
            f"✅ No items expiring within {days} days.",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    INGREDIENT_WHS = {"WDR12A", "WCS10A"}

    expiring.sort(key=lambda x: x["days_left"])
    by_item_proc = {}   # Processed Foods
    by_item_trd  = {}   # Trading (TR prefix)
    by_item_ing  = {}   # Ingredients & Spices (FA/Component warehouses)

    for r in expiring:
        k = r["item_no"]
        whs = r.get("whs_code", "").upper().strip()
        if whs in INGREDIENT_WHS:
            target = by_item_ing
        elif k.upper().startswith("TR"):
            target = by_item_trd
        else:
            target = by_item_proc
        if k not in target:
            target[k] = {"item_no": k, "desc": r["desc"], "total": 0.0, "min_days": r["days_left"], "min_exp": r["exp_date_obj"]}
        target[k]["total"] += r["in_stock"]
        if r["days_left"] < target[k]["min_days"]:
            target[k]["min_days"] = r["days_left"]
            target[k]["min_exp"] = r["exp_date_obj"]

    proc_items = sorted(by_item_proc.values(), key=lambda x: x["min_days"])
    trd_items  = sorted(by_item_trd.values(),  key=lambda x: x["min_days"])
    ing_items  = sorted(by_item_ing.values(),  key=lambda x: x["min_days"])
    await _send_expiring_split(update, proc_items, trd_items, ing_items, days)


def _expiry_emoji(days_left: int) -> str:
    if days_left < 0:   return "🔴"
    if days_left <= 7:  return "🔴"
    if days_left <= 30: return "🟠"
    return "🟡"


async def _send_expiring_list(update_or_query, items, days, show_all=False):
    """Send expiring items list. show_all=True skips truncation."""
    PREVIEW_LIMIT = 20
    header = f"⚠️ *Items Expiring Within {days} Days* ({len(items)} items)"
    lines = [header, ""]

    display_items = items if show_all else items[:PREVIEW_LIMIT]
    for item in display_items:
        d = item["min_days"]
        emoji = _expiry_emoji(d)
        exp_str = item["min_exp"].strftime("%m/%d/%Y")
        if d < 0:
            days_label = f"{abs(d)}d overdue"
        else:
            days_label = f"{d}d"
        lines.append(f"{emoji} *{item['desc']}* (`{item['item_no']}`)")  
        lines.append(f"  Qty: {int(item['total']):,} | Expires: {exp_str} ({days_label})")

    if not show_all and len(items) > PREVIEW_LIMIT:
        lines.append(f"")
        lines.append(f"_Showing {PREVIEW_LIMIT} of {len(items)} items_")

    lines.append(f"")
    lines.append(inv_source_footer())
    text = "\n".join(lines)

    MAX_LEN = 4000

    if not show_all and len(items) > PREVIEW_LIMIT:
        cb_data = f"exp_all|{days}"
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton(
            f"📋 Show All ({len(items)} items)", callback_data=cb_data
        )]])
        if hasattr(update_or_query, 'message') and update_or_query.message:
            await update_or_query.message.reply_text(text[:MAX_LEN], parse_mode=ParseMode.MARKDOWN, reply_markup=keyboard)
        else:
            await update_or_query.edit_message_text(text[:MAX_LEN], parse_mode=ParseMode.MARKDOWN, reply_markup=keyboard)
    else:
        # Split into chunks if too long
        chunks = []
        current = ""
        for line in lines:
            if len(current) + len(line) + 1 > MAX_LEN:
                chunks.append(current)
                current = line
            else:
                current = current + "\n" + line if current else line
        if current:
            chunks.append(current)

        if hasattr(update_or_query, 'message') and update_or_query.message:
            for chunk in chunks:
                await update_or_query.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
        else:
            # For callback query, edit first message then send rest as new messages
            await update_or_query.edit_message_text(chunks[0], parse_mode=ParseMode.MARKDOWN)
            for chunk in chunks[1:]:
                await update_or_query.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)


def _build_expiring_section(title: str, items: list, days: int, preview: int = 20, show_all: bool = False) -> tuple[str, bool]:
    """Build a section of expiring items. Returns (text, has_more)."""
    if not items:
        return f"{title}\n_None_\n", False
    display = items if show_all else items[:preview]
    lines = [title, ""]
    for item in display:
        d = item["min_days"]
        emoji = _expiry_emoji(d)
        exp_str = item["min_exp"].strftime("%m/%d/%Y")
        days_label = f"{abs(d)}d overdue" if d < 0 else f"{d}d"
        lines.append(f"{emoji} *{item['desc']}* (`{item['item_no']}`)")  
        lines.append(f"  Qty: {int(item['total']):,} | Expires: {exp_str} ({days_label})")
    has_more = not show_all and len(items) > preview
    if has_more:
        lines.append(f"_...and {len(items) - preview} more_")
    lines.append("")
    return "\n".join(lines), has_more


async def _send_expiring_split(update_obj, proc_items: list, trd_items: list, ing_items: list, days: int, show_all: bool = False):
    """Send expiring items in 3 sections: Processed Foods, Trading, Ingredients."""
    MAX_LEN = 4000
    PREVIEW = 20
    total = len(proc_items) + len(trd_items) + len(ing_items)

    header = f"⚠️ *Items Expiring Within {days} Days* ({total} items)\n"

    proc_text, proc_more = _build_expiring_section(
        f"🏭 *Processed Foods* ({len(proc_items)} items)", proc_items, days, PREVIEW, show_all)
    trd_text, trd_more  = _build_expiring_section(
        f"🛒 *Trading* ({len(trd_items)} items)", trd_items, days, PREVIEW, show_all)
    ing_text, ing_more  = _build_expiring_section(
        f"🧪 *Ingredients & Spices* ({len(ing_items)} items)", ing_items, days, PREVIEW, show_all)

    has_more = proc_more or trd_more or ing_more
    full_text = header + proc_text + trd_text + ing_text + inv_source_footer()

    # Helper to send chunked messages
    async def send_chunked(text, reply_markup=None):
        if len(text) <= MAX_LEN:
            if hasattr(update_obj, 'message') and update_obj.message:
                await update_obj.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
            else:
                await update_obj.edit_message_text(text, parse_mode=ParseMode.MARKDOWN)
        else:
            chunks, cur = [], ""
            for line in text.split("\n"):
                addition = ("\n" + line) if cur else line
                if len(cur) + len(addition) > MAX_LEN:
                    chunks.append(cur)
                    cur = line
                else:
                    cur += addition
            if cur:
                chunks.append(cur)
            first = True
            for chunk in chunks:
                if first and hasattr(update_obj, 'message') and update_obj.message:
                    await update_obj.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup if first else None)
                elif first:
                    await update_obj.edit_message_text(chunk, parse_mode=ParseMode.MARKDOWN)
                else:
                    if hasattr(update_obj, 'message') and update_obj.message:
                        await update_obj.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
                    else:
                        await update_obj.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
                first = False

    if has_more:
        cb_data = f"exp_all|{days}"
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton(
            f"📋 Show All ({total} items)", callback_data=cb_data
        )]])
        await send_chunked(full_text, reply_markup=keyboard)
    else:
        await send_chunked(full_text)


async def cb_search_item_check(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback: user tapped an item code button from search results — show item detail."""
    query = update.callback_query
    user = query.from_user
    chat = query.message.chat if query.message is not None else None
    if chat is not None:
        log_chat(chat.id, chat.type, chat.title)
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if chat is not None and is_whitelisted_group(chat.id):
                pass  # whitelisted group: fully open for all members
            elif ACCESS_MODE == "hard":
                await _send_blocked_notice(query, str(user.id))
                return
            elif ACCESS_MODE == "soft":
                await _send_blocked_notice(query, str(user.id))
                # soft: continue
    await query.answer()
    item_no = query.data.split("|", 1)[1] if "|" in query.data else ""
    if not item_no:
        return
    records = [r for r in store.inventory if r["item_no"].upper() == item_no.upper()]
    if not records:
        await query.message.reply_text(f"❌ Item {item_no} not found in current data.")
        return
    await _do_item_detail(query, records)


async def cb_check_show_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback: show all batches for a /check item."""
    query = update.callback_query
    user = query.from_user
    chat = query.message.chat if query.message is not None else None
    if chat is not None:
        log_chat(chat.id, chat.type, chat.title)
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if chat is not None and is_whitelisted_group(chat.id):
                pass  # whitelisted group: fully open for all members
            elif ACCESS_MODE == "hard":
                await _send_blocked_notice(query, str(user.id))
                return
            elif ACCESS_MODE == "soft":
                await _send_blocked_notice(query, str(user.id))
                # soft: continue
    await query.answer()
    item_no = query.data.split("|", 1)[1] if "|" in query.data else ""
    if not item_no:
        return

    records = [r for r in store.inventory if r["item_no"].upper() == item_no.upper()]
    if not records:
        await query.edit_message_text(f"❌ Item `{item_no}` not found in current data.", parse_mode=ParseMode.MARKDOWN)
        return

    today = datetime.now(PHT).date()
    desc = records[0]["desc"]
    total_qty = sum(r["in_stock"] for r in records)

    whs_totals: dict[str, float] = {}
    for r in records:
        if r["whs_name"]:
            whs_totals[r["whs_name"]] = whs_totals.get(r["whs_name"], 0.0) + r["in_stock"]
    whs_sorted = sorted(whs_totals.items(), key=lambda x: x[1], reverse=True)

    batch_rows: dict[tuple, dict] = {}
    for r in records:
        key = (r["batch"], r["whs_name"])
        if key not in batch_rows:
            batch_rows[key] = {"batch": r["batch"], "whs": r["whs_name"], "qty": 0.0, "exp_date": r["exp_date"]}
        batch_rows[key]["qty"] += r["in_stock"]

    exp_dates = [v["exp_date"] for v in batch_rows.values() if v["exp_date"] is not None]
    nearest_exp = min(exp_dates) if exp_dates else None

    def exp_tag(exp_date):
        if exp_date is None: return ""
        d = (exp_date - today).days
        if d < 0: return f"🔴 EXPIRED ({abs(d)}d ago)"
        elif d <= 30: return f"🟠 {d}d left"
        elif d <= 90: return f"🟡 {d}d left"
        else: return f"🟢 {d}d left"

    def parse_mfg(batch):
        digits = ''.join(c for c in batch if c.isdigit())
        if len(digits) >= 6:
            try: return datetime.strptime(digits[0:6], "%y%m%d").date()
            except ValueError: pass
        return None

    batch_list = sorted(batch_rows.values(), key=lambda x: (x["exp_date"] or date.max, x["whs"]))

    lines = [
        f"📦 *{desc}*",
        f"Item Code: `{item_no}`",
        f"Total Quantity: *{int(total_qty):,}*",
        f"Batches: {len(batch_rows)} | Warehouses: {len(whs_totals)}",
    ]
    if nearest_exp:
        d = (nearest_exp - today).days
        lines.append(f"{exp_tag(nearest_exp).split()[0]} Nearest Expiry: {nearest_exp.strftime('%m/%d/%Y')} ({abs(d)} days {'left' if d >= 0 else 'ago'})")
    lines.append("")
    lines.append("*Warehouse Breakdown:*")
    for whs, qty in whs_sorted:
        lines.append(f"• {whs}: {int(qty):,}")
    lines.append("")
    lines.append("*Batch Details (All):*")
    for b in batch_list:
        mfg = parse_mfg(b["batch"])
        exp = b["exp_date"]
        lines.append(f"• `{b['batch']}` | Qty: {int(b['qty']):,} | {b['whs']}")
        lines.append(f"  Mfg: {mfg.strftime('%m/%d/%Y') if mfg else '—'} | Exp: {exp.strftime('%m/%d/%Y') if exp else '—'} {exp_tag(exp)}")
    lines.append(f"\n{inv_source_footer()}")

    MAX_LEN = 4000
    full_text = "\n".join(lines)
    if len(full_text) <= MAX_LEN:
        await query.edit_message_text(full_text, parse_mode=ParseMode.MARKDOWN)
    else:
        # Edit first chunk, send rest as new messages
        chunks, cur = [], ""
        for line in lines:
            addition = ("\n" + line) if cur else line
            if len(cur) + len(addition) > MAX_LEN:
                chunks.append(cur)
                cur = line
            else:
                cur += addition
        if cur:
            chunks.append(cur)
        await query.edit_message_text(chunks[0], parse_mode=ParseMode.MARKDOWN)
        for chunk in chunks[1:]:
            await query.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)


async def cb_expiring_show_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback: expand expiring list to show all items (3 sections)."""
    query = update.callback_query
    user = query.from_user
    chat = query.message.chat if query.message is not None else None
    if chat is not None:
        log_chat(chat.id, chat.type, chat.title)
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if chat is not None and is_whitelisted_group(chat.id):
                pass  # whitelisted group: fully open for all members
            elif ACCESS_MODE == "hard":
                await _send_blocked_notice(query, str(user.id))
                return
            elif ACCESS_MODE == "soft":
                await _send_blocked_notice(query, str(user.id))
                # soft: continue
    await query.answer()
    parts = query.data.split("|")
    days = int(parts[1]) if len(parts) > 1 else 30

    INGREDIENT_WHS = {"WDR12A", "WCS10A"}
    today = datetime.now(PHT).date()
    expiring = []
    for r in store.inventory:
        if r["in_stock"] <= 0:
            continue
        exp_date = r.get("exp_date")
        if exp_date is None:
            continue
        days_left = (exp_date - today).days
        if days_left <= days:
            expiring.append({**r, "days_left": days_left, "exp_date_obj": exp_date})

    by_item_proc = {}
    by_item_trd  = {}
    by_item_ing  = {}

    for r in expiring:
        k = r["item_no"]
        whs = r.get("whs_code", "").upper().strip()
        if whs in INGREDIENT_WHS:
            target = by_item_ing
        elif k.upper().startswith("TR"):
            target = by_item_trd
        else:
            target = by_item_proc
        if k not in target:
            target[k] = {"item_no": k, "desc": r["desc"], "total": 0.0, "min_days": r["days_left"], "min_exp": r["exp_date_obj"]}
        target[k]["total"] += r["in_stock"]
        if r["days_left"] < target[k]["min_days"]:
            target[k]["min_days"] = r["days_left"]
            target[k]["min_exp"] = r["exp_date_obj"]

    proc_items = sorted(by_item_proc.values(), key=lambda x: x["min_days"])
    trd_items  = sorted(by_item_trd.values(),  key=lambda x: x["min_days"])
    ing_items  = sorted(by_item_ing.values(),  key=lambda x: x["min_days"])
    await _send_expiring_split(query, proc_items, trd_items, ing_items, days, show_all=True)


async def cmd_low(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_low
    threshold = LOW_STOCK_THRESHOLD
    if context.args:
        try:
            threshold = float(context.args[0])
        except ValueError:
            await update.message.reply_text("Usage: `/low [threshold]`\nExample: `/low 50`", parse_mode=ParseMode.MARKDOWN)
            return

    if not store.inventory:
        await update.message.reply_text("⚠️ Inventory data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    by_item = group_inventory_by_item(store.inventory)
    low_items = [(k, v) for k, v in by_item.items() if 0 < v["total"] <= threshold]
    low_items.sort(key=lambda x: x[1]["total"])

    if not low_items:
        await update.message.reply_text(f"✅ No items below {threshold:,.0f} units.", parse_mode=ParseMode.MARKDOWN)
        return

    lines = [f"⚠️ *Low Stock* (≤ {threshold:,.0f} units) — {len(low_items)} items\n"]
    for _, item in low_items[:20]:
        lines.append(f"• *{item['item_no']}* — {item['desc'][:40]}")
        lines.append(f"  Total: *{item['total']:,.0f}*")

    if len(low_items) > 20:
        lines.append(f"\n_Showing 20 of {len(low_items)} items._")

    lines.append(f"\n{inv_source_footer()}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_lowstock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_lowstock
    """/lowstock [keyword|all] [threshold] — show items below threshold qty, grouped by category."""
    DEFAULT_THRESHOLD = 500.0
    args = list(context.args) if context.args else []

    # Parse args: optional keyword and optional threshold
    keyword = ""
    threshold = DEFAULT_THRESHOLD
    if args:
        # Check if last arg is a number (threshold override)
        try:
            threshold = float(args[-1])
            keyword = " ".join(args[:-1]).strip().lower()
        except ValueError:
            keyword = " ".join(args).strip().lower()

    if not store.inventory:
        await update.message.reply_text("⚠️ Inventory data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    by_item = group_inventory_by_item(store.inventory)
    show_all_cats = (keyword in ("", "all"))

    # Filter items by keyword
    if show_all_cats:
        candidate_items = list(by_item.values())
    else:
        candidate_items = [
            v for v in by_item.values()
            if keyword in v["desc"].lower() or keyword in v["item_no"].lower()
        ]

    # Keep only items with stock > 0 and below threshold
    low_items = [v for v in candidate_items if 0 < v["total"] < threshold]
    low_items.sort(key=lambda x: x["total"], reverse=True)  # highest qty first

    if not low_items:
        label = "all items" if show_all_cats else f"*{keyword.upper()}* items"
        await update.message.reply_text(
            f"✅ No {label} below *{threshold:,.0f}* kg.",
            parse_mode=ParseMode.MARKDOWN
        )
        return

    def assign_category(desc: str) -> str:
        """Assign a display category to an item description."""
        desc_lower = desc.lower()
        for cat, kws in CATEGORY_KEYWORDS.items():
            if any(kw in desc_lower for kw in kws):
                return cat
        # Raw protein fallback
        for raw in ["beef", "pork", "chicken", "lamb", "veal", "duck", "fish", "shrimp", "prawn", "ground"]:
            if raw in desc_lower:
                return raw.upper()
        return "OTHER"

    if show_all_cats:
        # Group by category
        grouped: dict[str, list] = {}
        for item in low_items:
            cat = assign_category(item["desc"])
            grouped.setdefault(cat, []).append(item)

        header_line = f"⚠️ *Low Stock — All Categories* (< {threshold:,.0f} kg)\n{len(low_items)} item(s) below threshold\n"
        current_chunk = header_line
        first_msg = True

        for cat in sorted(grouped.keys()):
            items_in_cat = grouped[cat]
            section = f"\n*{cat}* ({len(items_in_cat)})\n"
            for item in items_in_cat:
                section += f"• `{item['item_no']}` {item['desc'][:45]}: *{item['total']:,.2f}* kg\n"

            if len(current_chunk) + len(section) > 3800:
                await update.message.reply_text(current_chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)
                first_msg = False
                current_chunk = section
            else:
                current_chunk += section

        current_chunk += f"\n{inv_source_footer()}"
        await update.message.reply_text(current_chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)

    else:
        # Single keyword — flat sorted list
        header_line = f"⚠️ *Low Stock — {keyword.upper()}* (< {threshold:,.0f} kg)\n{len(low_items)} item(s) below threshold\n"
        lines = [header_line]
        for item in low_items:
            lines.append(f"• `{item['item_no']}` {item['desc'][:45]}: *{item['total']:,.2f}* kg")
        lines.append(f"\n{inv_source_footer()}")

        current_chunk = ""
        first_msg = True
        for line in lines:
            if len(current_chunk) + len(line) + 1 > 3800:
                await update.message.reply_text(current_chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)
                first_msg = False
                current_chunk = line + "\n"
            else:
                current_chunk += line + "\n"
        if current_chunk.strip():
            await update.message.reply_text(current_chunk.rstrip(), parse_mode=ParseMode.MARKDOWN)


async def cmd_warehouse(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_warehouse
    query = " ".join(context.args).strip() if context.args else ""
    if not query:
        if not store.inventory:
            await update.message.reply_text("⚠️ Inventory data not loaded yet.", parse_mode=ParseMode.MARKDOWN)
            return
        whs_totals: dict[str, float] = {}
        whs_code_map: dict[str, str] = {}  # whs_name -> whs_code
        for r in store.inventory:
            if r["whs_name"]:
                whs_totals[r["whs_name"]] = whs_totals.get(r["whs_name"], 0.0) + r["in_stock"]
                if r["whs_code"] and r["whs_name"] not in whs_code_map:
                    whs_code_map[r["whs_name"]] = r["whs_code"]
        whs_list = sorted(whs_totals.items(), key=lambda x: x[1], reverse=True)
        lines = [f"🏭 *Warehouses* ({len(whs_list)} total)\n"]
        for whs, total in whs_list:
            code = whs_code_map.get(whs, "")
            code_str = f" `[{code}]`" if code else ""
            lines.append(f"• {whs}{code_str}: {total:,.0f}")
        lines.append(f"\n{inv_source_footer()}")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)
        return

    q = query.lower()
    matches = [r for r in store.inventory if q in r["whs_name"].lower() or q == r["whs_code"].lower()]
    if not matches:
        await update.message.reply_text(f"❌ No warehouse matching *{query}*.", parse_mode=ParseMode.MARKDOWN)
        return

    matched_whs = list(set(r["whs_name"] for r in matches))
    whs_label = ', '.join(sorted(matched_whs)[:2])
    # Build whs_code for the label
    whs_code_map: dict[str, str] = {}
    for r in matches:
        if r["whs_code"] and r["whs_name"] not in whs_code_map:
            whs_code_map[r["whs_name"]] = r["whs_code"]
    whs_code_str = "/".join(whs_code_map.get(w, "") for w in sorted(matched_whs)[:2] if whs_code_map.get(w, ""))
    whs_display = f"{whs_label} `[{whs_code_str}]`" if whs_code_str else whs_label

    # Count unique batches
    total_batches = len(set((r["batch"], r["whs_name"]) for r in matches if r["batch"]))
    total_qty = int(sum(r["in_stock"] for r in matches))

    by_item = group_inventory_by_item(matches)
    items = sorted(by_item.values(), key=lambda x: x["total"], reverse=True)
    items_with_stock = [i for i in items if i["total"] > 0]
    unique_items = len(items_with_stock)

    await _send_warehouse_detail(
        update, whs_label, unique_items, total_batches, total_qty, items_with_stock, show_all=False, whs_display=whs_display
    )


async def _send_warehouse_detail(update_or_query, whs_label, unique_items, total_batches, total_qty, items, show_all=False, whs_display=None):
    """Send warehouse detail. show_all=True sends full list without truncation."""
    PREVIEW_LIMIT = 20
    header_label = whs_display if whs_display else whs_label
    lines = [
        f"🏰 *{header_label}*",
        f"Unique Items: {unique_items} | Total Batches: {total_batches}",
        f"Total Quantity: *{total_qty:,}*",
        "",
        "*Items:*",
    ]
    display_items = items if show_all else items[:PREVIEW_LIMIT]
    for item in display_items:
        lines.append(f"• {item['desc']} (`{item['item_no']}`): {int(item['total']):,}")

    if not show_all and len(items) > PREVIEW_LIMIT:
        lines.append(f"\n_Showing {PREVIEW_LIMIT} of {len(items)} items_")

    lines.append(f"\n{inv_source_footer()}")

    if not show_all and len(items) > PREVIEW_LIMIT:
        # Preview mode: single message with keyboard button
        text = "\n".join(lines)
        cb_data = f"whs_all|{whs_label}|{unique_items}|{total_batches}|{total_qty}"
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton(
            f"📋 Show All ({len(items)} items)", callback_data=cb_data
        )]])
        try:
            if hasattr(update_or_query, 'message') and update_or_query.message:
                await update_or_query.message.reply_text(text[:4090], parse_mode=ParseMode.MARKDOWN, reply_markup=keyboard)
            else:
                await update_or_query.edit_message_text(text[:4090], parse_mode=ParseMode.MARKDOWN, reply_markup=keyboard)
        except Exception as e:
            logger.error(f"_send_warehouse_detail preview failed: {e}")
    else:
        # Show-all mode: chunk into multiple messages
        MAX_LEN = 3500
        chunks, cur = [], ""
        for line in lines:
            addition = ("\n" + line) if cur else line
            if len(cur) + len(addition) > MAX_LEN:
                chunks.append(cur)
                cur = line
            else:
                cur += addition
        if cur:
            chunks.append(cur)
        if not chunks:
            return
        is_callback = hasattr(update_or_query, 'edit_message_text')
        try:
            if is_callback:
                await update_or_query.edit_message_text(chunks[0][:4090], parse_mode=ParseMode.MARKDOWN)
            else:
                await update_or_query.message.reply_text(chunks[0][:4090], parse_mode=ParseMode.MARKDOWN)
        except Exception as e:
            logger.error(f"_send_warehouse_detail chunk 0 failed: {e}")
            try:
                await update_or_query.message.reply_text(chunks[0][:4090])
            except Exception:
                pass
        for i, chunk in enumerate(chunks[1:], start=1):
            try:
                await update_or_query.message.reply_text(chunk[:4090], parse_mode=ParseMode.MARKDOWN)
            except Exception as e:
                logger.error(f"_send_warehouse_detail chunk {i} failed: {e}")
                try:
                    await update_or_query.message.reply_text(chunk[:4090])
                except Exception:
                    pass


async def cb_warehouse_show_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback: expand warehouse list to show all items."""
    query = update.callback_query
    user = query.from_user
    chat = query.message.chat if query.message is not None else None
    if chat is not None:
        log_chat(chat.id, chat.type, chat.title)
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if chat is not None and is_whitelisted_group(chat.id):
                pass  # whitelisted group: fully open for all members
            elif ACCESS_MODE == "hard":
                await _send_blocked_notice(query, str(user.id))
                return
            elif ACCESS_MODE == "soft":
                await _send_blocked_notice(query, str(user.id))
                # soft: continue
    await query.answer()
    data = query.data  # whs_all|whs_label|unique|batches|qty
    parts = data.split("|")
    if len(parts) < 5:
        await query.edit_message_text("⚠️ Could not load full list.", parse_mode=ParseMode.MARKDOWN)
        return

    whs_label = parts[1]
    unique_items = int(parts[2])
    total_batches = int(parts[3])
    total_qty = int(parts[4])

    # Re-fetch items for this warehouse
    q = whs_label.lower()
    # Match any warehouse that is part of the label
    whs_names = [n for n in whs_label.split(', ')]
    matches = [r for r in store.inventory if any(n.lower() in r["whs_name"].lower() or r["whs_name"].lower() in n.lower() for n in whs_names)]
    by_item = group_inventory_by_item(matches)
    items = sorted(by_item.values(), key=lambda x: x["total"], reverse=True)
    items_with_stock = [i for i in items if i["total"] > 0]

    await _send_warehouse_detail(
        query, whs_label, unique_items, total_batches, total_qty, items_with_stock, show_all=True
    )


async def cmd_top(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_top
    if not store.inventory:
        await update.message.reply_text("⚠️ Inventory data not loaded yet.", parse_mode=ParseMode.MARKDOWN)
        return

    by_item = group_inventory_by_item(store.inventory)
    top_items = sorted(by_item.values(), key=lambda x: x["total"], reverse=True)[:20]

    lines = ["🏆 *Top 20 Items by Quantity*\n"]
    for i, item in enumerate(top_items, 1):
        lines.append(f"{i}. *{item['item_no']}* — {item['desc'][:35]}")
        lines.append(f"   {item['total']:,.0f} units")

    lines.append(f"\n{inv_source_footer()}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


def _exp_tag_full(batches, today):
    """Return expiry tag always showing the date (not just near-expiry items)."""
    dates = [b["exp_date"] for b in batches if b.get("exp_date")]
    if not dates:
        return ""
    nearest = min(dates)
    d = (nearest - today).days
    if d < 0:
        return f" 🔴 EXPIRED (exp {nearest.strftime('%m/%d/%y')})"
    elif d <= 30:
        return f" 🟠 exp {nearest.strftime('%m/%d/%y')} ({d}d)"
    elif d <= 90:
        return f" 🟡 exp {nearest.strftime('%m/%d/%y')} ({d}d)"
    else:
        return f" exp {nearest.strftime('%m/%d/%y')}"


async def cmd_components(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_components
    """/components [keyword] — Component Warehouse stock for PRD use."""
    keyword = " ".join(context.args).strip().lower() if context.args else ""

    if not store.component_inventory:
        await update.message.reply_text(
            "⚠️ Component Warehouse data not loaded yet. Try `/refresh`.",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    records = store.component_inventory
    if keyword:
        records = [r for r in records if keyword in r["desc"].lower() or keyword in r["item_no"].lower()]
        if not records:
            await update.message.reply_text(
                f"❌ No components matching *{keyword}*.",
                parse_mode=ParseMode.MARKDOWN,
            )
            return

    by_item = group_inventory_by_item(records)
    items = sorted(by_item.values(), key=lambda x: x["total"], reverse=True)
    items_with_stock = [i for i in items if i["total"] > 0]
    items_zero = [i for i in items if i["total"] <= 0]

    today = datetime.now(PHT).date()
    total_qty = sum(i["total"] for i in items_with_stock)
    total_items = len(items)
    PREVIEW_LIMIT = 20

    header = (
        f"🧪 *Component Warehouse*"
        + (f" — search: _{keyword}_" if keyword else "")
        + f"\n{total_items} item(s) | Total qty: *{total_qty:,.0f}*\n"
    )

    lines = [header]
    def _safe_desc(s):
        """Strip Markdown special chars from item descriptions to prevent parse errors."""
        return re.sub(r'[*_`\[\]()]', '', s)

    display_items = items_with_stock[:PREVIEW_LIMIT]
    for item in display_items:
        tag = _exp_tag_full(item["batches"], today)
        safe = _safe_desc(item['desc'][:45])
        lines.append(f"• `{item['item_no']}` {safe}: *{item['total']:,.0f}*{tag}")

    has_more = len(items_with_stock) > PREVIEW_LIMIT
    if has_more:
        lines.append(f"\n_Showing {PREVIEW_LIMIT} of {len(items_with_stock)} items with stock._")
    if items_zero:
        lines.append(f"_({len(items_zero)} item(s) with zero stock not shown)_")
    lines.append(f"\n{inv_source_footer()}")

    text = "\n".join(lines)
    if has_more:
        kw_safe = keyword.replace("|", "")  # sanitise for callback data
        cb_data = f"comp_all|{kw_safe}"
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton(
            f"📋 Show All ({len(items_with_stock)} items)", callback_data=cb_data
        )]])
        try:
            await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=keyboard)
        except Exception as e:
            logger.error(f"cmd_components preview failed (markdown): {e}")
            await update.message.reply_text(text, reply_markup=keyboard)
    else:
        try:
            await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
        except Exception as e:
            logger.error(f"cmd_components display failed (markdown): {e}")
            await update.message.reply_text(text)


async def cmd_components_expiring(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
# placeholder for cmd_components_expiring
    """/components_expiring [days] — Component Warehouse items expiring within N days."""
    raw = context.args[0] if context.args else "30"
    try:
        days = int(raw)
    except ValueError:
        await update.message.reply_text(
            "⚠️ Usage: `/components_expiring [days]` — e.g. `/components_expiring 60`",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    if not store.component_inventory:
        await update.message.reply_text(
            "⚠️ Component Warehouse data not loaded yet. Try `/refresh`.",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    today = datetime.now(PHT).date()
    cutoff = today + timedelta(days=days)

    by_item = group_inventory_by_item(store.component_inventory)

    expiring_items = []
    for item in by_item.values():
        dates = [b["exp_date"] for b in item["batches"] if b.get("exp_date")]
        if not dates:
            continue
        nearest = min(dates)
        if nearest <= cutoff:
            expiring_items.append((nearest, item))

    expiring_items.sort(key=lambda x: x[0])  # sort by nearest expiry first

    if not expiring_items:
        label = "already expired" if days == 0 else f"expiring within {days} day(s)"
        await update.message.reply_text(
            f"✅ No components {label}.",
        )
        return

    expired = [(d, i) for d, i in expiring_items if d < today]
    soon = [(d, i) for d, i in expiring_items if d >= today]

    label = "already expired" if days == 0 else f"expiring within {days} day(s)"
    header = (
        f"🧪 *Component Warehouse — {label.title()}*\n"
        f"{len(expiring_items)} item(s) found\n"
    )
    lines = [header]

    if expired:
        lines.append("*🔴 Already Expired:*")
        for d, item in expired:
            d_str = d.strftime('%m/%d/%y')
            lines.append(f"• `{item['item_no']}` {item['desc'][:40]}: *{item['total']:,.0f}* (exp {d_str})")
        lines.append("")

    if soon:
        lines.append(f"*🟠 Expiring by {cutoff.strftime('%m/%d/%y')}:*")
        for d, item in soon:
            d_days = (d - today).days
            d_str = d.strftime('%m/%d/%y')
            emoji = "🟠" if d_days <= 30 else "🟡"
            lines.append(f"• {emoji} `{item['item_no']}` {item['desc'][:40]}: *{item['total']:,.0f}* (exp {d_str}, {d_days}d)")

    lines.append(f"\n{inv_source_footer()}")

    MAX_LEN = 4000
    full_text = "\n".join(lines)
    if len(full_text) <= MAX_LEN:
        await update.message.reply_text(full_text, parse_mode=ParseMode.MARKDOWN)
    else:
        chunks, cur = [], ""
        for line in lines:
            addition = ("\n" + line) if cur else line
            if len(cur) + len(addition) > MAX_LEN:
                chunks.append(cur)
                cur = line
            else:
                cur += addition
        if cur:
            chunks.append(cur)
        for chunk in chunks:
            await update.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)


async def cb_components_show_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback: expand component list to show all items."""
    query = update.callback_query
    user = query.from_user
    chat = query.message.chat if query.message is not None else None
    if chat is not None:
        log_chat(chat.id, chat.type, chat.title)
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if chat is not None and is_whitelisted_group(chat.id):
                pass  # whitelisted group: fully open for all members
            elif ACCESS_MODE == "hard":
                await _send_blocked_notice(query, str(user.id))
                return
            elif ACCESS_MODE == "soft":
                await _send_blocked_notice(query, str(user.id))
                # soft: continue
    # Answer the callback immediately to dismiss the loading spinner
    await query.answer()
    parts = query.data.split("|", 1)
    keyword = parts[1].lower() if len(parts) > 1 else ""

    # Schedule the heavy work as a background task so the webhook
    # returns quickly and doesn't hit the 60s timeout
    async def _send_all():
        try:
            records = store.component_inventory
            if keyword:
                records = [r for r in records if keyword in r["desc"].lower() or keyword in r["item_no"].lower()]

            today = datetime.now(PHT).date()
            by_item = group_inventory_by_item(records)
            items = sorted(by_item.values(), key=lambda x: x["total"], reverse=True)
            items_with_stock = [i for i in items if i["total"] > 0]

            total_qty = sum(i["total"] for i in items_with_stock)
            header = (
                f"🧪 *Component Warehouse — All Items*"
                + (f" (_{keyword}_)" if keyword else "")
                + f"\n{len(items_with_stock)} item(s) | Total qty: *{total_qty:,.0f}*\n"
            )

            def _safe_desc(s):
                return re.sub(r'[*_`\[\]()]', '', s)

            lines = [header]
            for item in items_with_stock:
                tag = _exp_tag_full(item["batches"], today)
                safe = _safe_desc(item['desc'][:45])
                lines.append(f"• `{item['item_no']}` {safe}: *{item['total']:,.0f}*{tag}")
            lines.append(f"\n{inv_source_footer()}")

            MAX_LEN = 3500  # conservative limit well below Telegram's 4096
            chunks, cur = [], ""
            for line in lines:
                addition = ("\n" + line) if cur else line
                if len(cur) + len(addition) > MAX_LEN:
                    chunks.append(cur)
                    cur = line
                else:
                    cur += addition
            if cur:
                chunks.append(cur)

            if not chunks:
                await query.message.reply_text("⚠️ No items found.")
                return

            # Edit the original preview message with the first chunk
            try:
                await query.edit_message_text(chunks[0][:4090], parse_mode=ParseMode.MARKDOWN)
            except Exception as e:
                logger.error(f"cb_components_show_all edit chunk 0 failed: {e}")
                try:
                    await query.message.reply_text(chunks[0][:4090], parse_mode=ParseMode.MARKDOWN)
                except Exception:
                    await query.message.reply_text(chunks[0][:4090])
            # Send remaining chunks as follow-up messages
            for i, chunk in enumerate(chunks[1:], start=1):
                try:
                    await query.message.reply_text(chunk[:4090], parse_mode=ParseMode.MARKDOWN)
                except Exception as e:
                    logger.error(f"cb_components_show_all reply chunk {i} failed: {e}")
                    await query.message.reply_text(chunk[:4090])
        except Exception as e:
            logger.error(f"cb_components_show_all _send_all failed: {e}", exc_info=True)
            try:
                await query.message.reply_text(f"⚠️ Failed to load full list: {e}")
            except Exception:
                pass

    # Schedule as background task — does not block the webhook thread
    context.application.create_task(_send_all())


# ──────────────────────────────────────────────────────────────────────────────
# Command handlers — AR
# ──────────────────────────────────────────────────────────────────────────────
async def cmd_ar_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    total = sum(r["balance"] for r in store.ar_rows)
    clients = len(set(r["card_name"] for r in store.ar_rows))
    si_count = len(store.ar_rows)

    # Aging buckets matching old bot: Current, 1-7d, 8-30d, 31-60d, 61-90d, 90+d
    buckets = {"Current": 0.0, "1-7 days": 0.0, "8-30 days": 0.0, "31-60 days": 0.0, "61-90 days": 0.0, "90+ days": 0.0}
    for r in store.ar_rows:
        d = r["days_due"]
        if d <= 0:
            buckets["Current"] += r["balance"]
        elif d <= 7:
            buckets["1-7 days"] += r["balance"]
        elif d <= 30:
            buckets["8-30 days"] += r["balance"]
        elif d <= 60:
            buckets["31-60 days"] += r["balance"]
        elif d <= 90:
            buckets["61-90 days"] += r["balance"]
        else:
            buckets["90+ days"] += r["balance"]

    # Top 5 clients by total balance
    by_client: dict[str, float] = {}
    for r in store.ar_rows:
        by_client[r["card_name"]] = by_client.get(r["card_name"], 0.0) + r["balance"]
    top_clients = sorted(by_client.items(), key=lambda x: x[1], reverse=True)[:5]

    # Top 5 areas by total balance (using SEGMENT field from BP_ByCustomer_Segment)
    by_area: dict[str, float] = {}
    for r in store.ar_rows:
        area = r.get("area", "").strip() or "UNCLASSIFIED"
        by_area[area] = by_area.get(area, 0.0) + r["balance"]
    top_areas = sorted(by_area.items(), key=lambda x: x[1], reverse=True)[:5]

    ts = store.last_refresh.strftime("%m/%d/%Y") if store.last_refresh else "—"

    lines = [
        f"📊 AR Executive Summary",
        f"As of {ts} PHT",
        f"",
        f"Total Outstanding: {fmt_peso(total)}",
        f"Clients: {clients} | Open SIs: {si_count}",
        f"",
        f"Aging Breakdown:",
    ]
    for bucket, amount in buckets.items():
        pct = (amount / total * 100) if total > 0 else 0
        lines.append(f"  {bucket}: {fmt_peso(amount)} ({pct:.1f}%)")

    lines.append(f"")
    lines.append(f"Top 5 Clients by Balance:")
    for i, (name, bal) in enumerate(top_clients, 1):
        lines.append(f"  {i}. {name}: {fmt_peso(bal)}")

    lines.append(f"")
    lines.append(f"Top 5 Areas by Balance:")
    for area, bal in top_areas:
        lines.append(f"  • {area}: {fmt_peso(bal)}")

    lines.append(f"")
    lines.append(ar_source_footer())
    await update.message.reply_text("\n".join(lines))


async def cmd_ar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    query = " ".join(context.args).strip() if context.args else ""
    if not query:
        await update.message.reply_text("Usage: `/ar <client name>`\nExample: `/ar jollibee`", parse_mode=ParseMode.MARKDOWN)
        return
    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    q = query.lower()
    # Exact code match takes priority (e.g. /client CMNL00013)
    code_matches = [r for r in store.ar_rows if r["card_code"].lower() == q]
    if code_matches:
        matches = code_matches
    else:
        matches = [r for r in store.ar_rows if q in r["card_name"].lower() or q in r["card_code"].lower()]
    if not matches:
        all_clients = list(set(r["card_name"] for r in store.ar_rows))
        close = difflib.get_close_matches(query, all_clients, n=3, cutoff=0.5)
        if close:
            await update.message.reply_text(
                f"❓ No exact match for *{query}*. Did you mean:\n" +
                "\n".join(f"• {c}" for c in close),
                parse_mode=ParseMode.MARKDOWN,
            )
        else:
            await update.message.reply_text(f"❌ No AR records for *{query}*.", parse_mode=ParseMode.MARKDOWN)
        return

    by_client: dict[str, dict] = {}
    for r in matches:
        k = r["card_name"]
        if k not in by_client:
            by_client[k] = {
                "name": k,
                "code": r["card_code"],
                "agent": r["agent"],
                "terms": r["terms"],
                "segment": r.get("area", "") or "",
                "total": 0.0,
                "sis": [],
            }
        by_client[k]["total"] += r["balance"]
        by_client[k]["sis"].append(r)

    if len(by_client) == 1:
        # Rich single-client detail view
        await _do_ar_client_detail(update, list(by_client.values())[0])
    else:
        # Compact multi-client list
        clients_sorted = sorted(by_client.values(), key=lambda x: x["total"], reverse=True)
        lines = [f'🔍 AR Search: "{query}" — {len(by_client)} clients found\n']
        for c in clients_sorted:
            si_count = len(c["sis"])
            lines.append(f"\u2022 <code>{c['code']}</code> {c['name']}: {fmt_peso(c['total'])} ({si_count} SIs)")
        lines.append(f"\nTip: Use /client [code] (e.g. CMNL00013) for direct lookup, or /client [exact name] for full details.")
        lines.append(ar_source_footer())
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


async def _do_ar_client_detail(update, client: dict):
    """Rich single-client AR detail: header, aging breakdown, SI details with overdue tags."""
    today = datetime.now(PHT).date()
    sis = client["sis"]
    total = client["total"]
    si_count = len(sis)

    # Aging buckets
    buckets = {"Current": 0.0, "1-7 days": 0.0, "8-30 days": 0.0, "31-60 days": 0.0, "61-90 days": 0.0, "90+ days": 0.0}
    for si in sis:
        d = si["days_due"]
        if d <= 0:
            buckets["Current"] += si["balance"]
        elif d <= 7:
            buckets["1-7 days"] += si["balance"]
        elif d <= 30:
            buckets["8-30 days"] += si["balance"]
        elif d <= 60:
            buckets["31-60 days"] += si["balance"]
        elif d <= 90:
            buckets["61-90 days"] += si["balance"]
        else:
            buckets["90+ days"] += si["balance"]

    def overdue_tag(days_due: int) -> str:
        if days_due <= 0:
            return "🟢 current"
        elif days_due <= 7:
            return f"🟡 {days_due}d overdue"
        elif days_due <= 30:
            return f"🟠 {days_due}d overdue"
        elif days_due <= 60:
            return f"🔴 {days_due}d overdue"
        else:
            return f"🔴 {days_due}d overdue"

    lines = [
        f"🗂 {client['name']}",
        f"Code: <code>{client['code']}</code> | Agent: {client['agent'] or '—'}",
        f"Terms: {client['terms'] or '—'} | Area: {client.get('segment', '') or '—'}",
        f"Total Outstanding: {fmt_peso(total)} ({si_count} SIs)",
        "",
        "Aging Breakdown:",
    ]
    for bucket, amount in buckets.items():
        if amount > 0:
            lines.append(f"  • {bucket}: {fmt_peso(amount)}")

    lines.append("")
    lines.append("SI Details:")
    sis_sorted = sorted(sis, key=lambda x: x["days_due"], reverse=True)
    for si in sis_sorted:
        bir = si.get("bir_si", "") or ""
        si_label = f"SI#{bir}" if bir else f"SI {si['si_number']}"
        si_date_str = fmt_date(si.get("si_date"))
        due_date_str = fmt_date(si.get("due_date"))
        tag = overdue_tag(si["days_due"])
        lines.append(f"  • {si_label} | {si_date_str} | Due: {due_date_str}")
        lines.append(f"    Balance: {fmt_peso(si['balance'])} | {tag}")

    lines.append(f"\n{ar_source_footer()}")
    full_text = "\n".join(lines)
    MAX_LEN = 4000
    if len(full_text) <= MAX_LEN:
        await update.message.reply_text(full_text, parse_mode=ParseMode.HTML)
    else:
        chunks, current = [], ""
        for line in lines:
            if len(current) + len(line) + 1 > MAX_LEN:
                chunks.append(current.rstrip())
                current = line + "\n"
            else:
                current += line + "\n"
        if current.strip():
            chunks.append(current.rstrip())
        for chunk in chunks:
            await update.message.reply_text(chunk, parse_mode=ParseMode.HTML)


# /client is an alias for /ar
async def cmd_client(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    await cmd_ar(update, context)


async def cmd_aging(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    buckets = {"Current (≤0d)": 0.0, "1-30d": 0.0, "31-60d": 0.0, "61-90d": 0.0, "90+d": 0.0}
    bucket_counts = {k: 0 for k in buckets}

    for r in store.ar_rows:
        d = r["days_due"]
        if d <= 0:
            buckets["Current (≤0d)"] += r["balance"]
            bucket_counts["Current (≤0d)"] += 1
        elif d <= 30:
            buckets["1-30d"] += r["balance"]
            bucket_counts["1-30d"] += 1
        elif d <= 60:
            buckets["31-60d"] += r["balance"]
            bucket_counts["31-60d"] += 1
        elif d <= 90:
            buckets["61-90d"] += r["balance"]
            bucket_counts["61-90d"] += 1
        else:
            buckets["90+d"] += r["balance"]
            bucket_counts["90+d"] += 1

    total = sum(buckets.values())
    lines = [f"📊 *AR Aging Summary*\n"]
    for bucket, amount in buckets.items():
        pct = (amount / total * 100) if total > 0 else 0
        lines.append(f"• *{bucket}*: {fmt_peso(amount)} ({pct:.1f}%) — {bucket_counts[bucket]} SIs")

    lines.append(f"\n💰 *Total Outstanding: {fmt_peso(total)}*")
    lines.append(ar_source_footer())
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_overdue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """Top overdue clients. Optional days threshold (default: any overdue)."""
    min_days = 1
    if context.args:
        try:
            min_days = int(context.args[0])
        except ValueError:
            await update.message.reply_text("Usage: `/overdue [days]`\nExample: `/overdue 30`", parse_mode=ParseMode.MARKDOWN)
            return

    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    overdue = [r for r in store.ar_rows if r["days_due"] >= min_days]
    by_client: dict[str, dict] = {}
    for r in overdue:
        k = r["card_name"]
        if k not in by_client:
            by_client[k] = {"name": k, "total": 0.0, "max_days": 0, "count": 0}
        by_client[k]["total"] += r["balance"]
        by_client[k]["max_days"] = max(by_client[k]["max_days"], r["days_due"])
        by_client[k]["count"] += 1

    top = sorted(by_client.values(), key=lambda x: x["total"], reverse=True)[:15]
    total_overdue = sum(r["balance"] for r in overdue)

    label = f"overdue ≥{min_days}d" if min_days > 1 else "overdue"
    lines = [f"⚠️ Top Overdue Clients ({len(by_client)} {label})\n"]
    for i, c in enumerate(top, 1):
        lines.append(f"{i}. {c['name']}")
        lines.append(f"   {fmt_peso(c['total'])} | Max: {c['max_days']}d | {c['count']} SIs")

    lines.append(f"\n💰 Total Overdue: {fmt_peso(total_overdue)}")
    lines.append(ar_source_footer())
    await update.message.reply_text("\n".join(lines))



EMPLOYEE_AREA_KEYWORDS = ("employee", "cdu clients")


async def cmd_top30overdue(update, context):
    if not await _access_gate(update, context):
        return
    """Top 20 clients with AR overdue 30+ days, grouped by area (excluding employee accounts)."""
    from telegram.constants import ParseMode
    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    MIN_DAYS = 30

    # Aggregate overdue 30+ days per client, excluding employee areas
    by_client = {}
    for r in store.ar_rows:
        if r["days_due"] < MIN_DAYS:
            continue
        area = r.get("area", "").strip() or "UNCLASSIFIED"
        if any(kw in area.lower() for kw in EMPLOYEE_AREA_KEYWORDS):
            continue
        k = r["card_name"]
        if k not in by_client:
            by_client[k] = {
                "name": k,
                "code": r["card_code"],
                "area": area,
                "overdue_total": 0.0,
                "max_days": 0,
                "si_count": 0,
            }
        by_client[k]["overdue_total"] += r["balance"]
        by_client[k]["max_days"] = max(by_client[k]["max_days"], r["days_due"])
        by_client[k]["si_count"] += 1

    if not by_client:
        await update.message.reply_text(f"✅ No clients with AR overdue {MIN_DAYS}+ days.")
        return

    # Pick top 20 by overdue balance
    top20 = sorted(by_client.values(), key=lambda x: x["overdue_total"], reverse=True)[:20]

    # Group top 20 by area
    by_area = {}
    for c in top20:
        by_area.setdefault(c["area"], []).append(c)

    # Sort areas by their total overdue balance descending
    area_order = sorted(by_area.keys(), key=lambda a: sum(c["overdue_total"] for c in by_area[a]), reverse=True)

    grand_total = sum(c["overdue_total"] for c in top20)
    total_clients_overdue = len(by_client)

    lines = [f"🔴 <b>Top 20 Clients \u2014 {MIN_DAYS}+ Days Overdue</b>\n({total_clients_overdue} clients total overdue {MIN_DAYS}+ days)\n"]
    rank = 1
    for area in area_order:
        clients = by_area[area]
        area_total = sum(c["overdue_total"] for c in clients)
        lines.append(f"\n━━ {area} ━━  <i>{fmt_peso(area_total)}</i>")
        for c in clients:
            lines.append(f"{rank}. <code>{c['code']}</code> {c['name']}")
            lines.append(f"   {fmt_peso(c['overdue_total'])} | Max: {c['max_days']}d overdue | {c['si_count']} SIs")
            rank += 1

    lines.append(f"\n💰 <b>Grand Total Overdue {MIN_DAYS}+ days: {fmt_peso(grand_total)}</b>")
    lines.append(ar_source_footer())

    text = "\n".join(lines)
    if len(text) <= 4096:
        await update.message.reply_text(text, parse_mode="HTML")
    else:
        chunks = []
        current = ""
        for line in lines:
            if len(current) + len(line) + 1 > 4000:
                chunks.append(current)
                current = line
            else:
                current = current + "\n" + line if current else line
        if current:
            chunks.append(current)
        for chunk in chunks:
            await update.message.reply_text(chunk, parse_mode="HTML")

async def cmd_area(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """Receivables grouped by area (SEGMENT field). /area [name] to filter."""
    area_filter = " ".join(context.args).strip().lower() if context.args else ""

    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    rows = store.ar_rows

    if area_filter:
        # Filter rows matching the area keyword
        matched_rows = [r for r in rows if area_filter in r.get("area", "").lower()]
        if not matched_rows:
            await update.message.reply_text(f"❌ No AR records for area *{area_filter.upper()}*.", parse_mode=ParseMode.MARKDOWN)
            return

        # Find distinct matching segment names
        matched_segments = sorted(set(r.get("area", "").strip() for r in matched_rows if r.get("area", "").strip()))

        # Check for exact match first (case-insensitive)
        exact_matches = [s for s in matched_segments if s.lower() == area_filter.lower()]
        if exact_matches:
            matched_segments = exact_matches
            matched_rows = [r for r in matched_rows if r.get("area", "").strip().lower() == area_filter.lower()]

        if len(matched_segments) > 1:
            # Disambiguation: multiple segments match — show summary list
            by_seg: dict[str, dict] = {}
            for r in matched_rows:
                seg = r.get("area", "").strip() or "UNCLASSIFIED"
                if seg not in by_seg:
                    by_seg[seg] = {"total": 0.0, "clients": set(), "count": 0}
                by_seg[seg]["total"] += r["balance"]
                by_seg[seg]["clients"].add(r["card_name"])
                by_seg[seg]["count"] += 1
            lines = [f"📍 Areas matching '{area_filter.upper()}' ({len(matched_segments)} found)\n"]
            for seg in sorted(by_seg.keys(), key=lambda s: by_seg[s]["total"], reverse=True):
                d = by_seg[seg]
                lines.append(f"• {seg}: {fmt_peso(d['total'])} — {len(d['clients'])} clients, {d['count']} SIs")
            lines.append(f"\nUse /area <full name> for details")
            lines.append(f"e.g. /area {matched_segments[0]}")
            lines.append(f"\n{ar_source_footer()}")
            await update.message.reply_text("\n".join(lines))
            return

        # Exactly one segment matches — show full detail view
        seg_name = matched_segments[0]
        seg_rows = [r for r in matched_rows if r.get("area", "").strip() == seg_name]
        total = sum(r["balance"] for r in seg_rows)
        clients = set(r["card_name"] for r in seg_rows)

        # Aging buckets
        aging_current = sum(r["balance"] for r in seg_rows if r["days_due"] <= 0)
        aging_31_60   = sum(r["balance"] for r in seg_rows if 1 <= r["days_due"] <= 30)
        aging_61_90   = sum(r["balance"] for r in seg_rows if 31 <= r["days_due"] <= 60)
        aging_91_120  = sum(r["balance"] for r in seg_rows if 61 <= r["days_due"] <= 90)
        aging_121plus = sum(r["balance"] for r in seg_rows if r["days_due"] > 90)
        total_overdue = aging_31_60 + aging_61_90 + aging_91_120 + aging_121plus

        client_totals: dict[str, float] = {}
        client_max_days: dict[str, int] = {}
        for r in seg_rows:
            client_totals[r["card_name"]] = client_totals.get(r["card_name"], 0.0) + r["balance"]
            client_max_days[r["card_name"]] = max(client_max_days.get(r["card_name"], 0), r["days_due"])

        lines = [f"📍 Area: {seg_name}\n"]
        lines.append(f"Total Balance: {fmt_peso(total)}")
        lines.append(f"Clients: {len(clients)} | SIs: {len(seg_rows)}")
        lines.append("")
        lines.append("Aging Summary:")
        lines.append(f"🟢 Current (not yet due):  {fmt_peso(aging_current)}")
        lines.append(f"🟡 1–30 days overdue:      {fmt_peso(aging_31_60)}")
        lines.append(f"🟠 31–60 days overdue:     {fmt_peso(aging_61_90)}")
        lines.append(f"🔴 61–90 days overdue:     {fmt_peso(aging_91_120)}")
        lines.append(f"🔴 91+ days overdue:       {fmt_peso(aging_121plus)}")
        lines.append(f"\n⚠️ Total Overdue: {fmt_peso(total_overdue)}")
        lines.append("")
        lines.append("Clients:")
        for name, bal in sorted(client_totals.items(), key=lambda x: x[1], reverse=True)[:20]:
            days = client_max_days.get(name, 0)
            days_str = f" — {days}d overdue" if days > 0 else " — current"
            lines.append(f"• {name}: {fmt_peso(bal)}{days_str}")
        if len(client_totals) > 20:
            lines.append(f"...and {len(client_totals) - 20} more clients")
        lines.append(f"\n{ar_source_footer()}")
        full_text = "\n".join(lines)
        MAX_LEN = 4000
        if len(full_text) <= MAX_LEN:
            await update.message.reply_text(full_text)
        else:
            chunks, current = [], ""
            for line in lines:
                if len(current) + len(line) + 1 > MAX_LEN:
                    chunks.append(current.rstrip())
                    current = line + "\n"
                else:
                    current += line + "\n"
            if current.strip():
                chunks.append(current.rstrip())
            for chunk in chunks:
                await update.message.reply_text(chunk)
        return

    # No filter — show all areas summary
    by_area: dict[str, dict] = {}
    for r in rows:
        area = r.get("area", "").strip() or "UNCLASSIFIED"
        if area not in by_area:
            by_area[area] = {"name": area, "total": 0.0, "overdue": 0.0, "clients": set(), "count": 0}
        by_area[area]["total"] += r["balance"]
        if r["days_due"] > 0:
            by_area[area]["overdue"] += r["balance"]
        by_area[area]["clients"].add(r["card_name"])
        by_area[area]["count"] += 1

    areas = sorted(by_area.values(), key=lambda x: x["total"], reverse=True)
    grand_total = sum(a["total"] for a in areas)

    lines = [f"📍 Receivables by Area ({len(areas)} areas)\n"]
    for a in areas:
        pct = (a["total"] / grand_total * 100) if grand_total > 0 else 0
        lines.append(f"• {a['name']}: {fmt_peso(a['total'])} ({pct:.1f}%) — {len(a['clients'])} clients")
    lines.append(f"\n💰 Grand Total: {fmt_peso(grand_total)}")
    lines.append(f"\n{ar_source_footer()}")
    await update.message.reply_text("\n".join(lines))


async def cmd_agent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """List clients under a specific agent."""
    agent_filter = " ".join(context.args).strip().lower() if context.args else ""
    if not agent_filter:
        await update.message.reply_text("Usage: `/agent <name>`\nExample: `/agent juan`", parse_mode=ParseMode.MARKDOWN)
        return

    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    matches = [r for r in store.ar_rows if agent_filter in r["agent"].lower()]
    if not matches:
        all_agents = list(set(r["agent"] for r in store.ar_rows if r["agent"]))
        close = difflib.get_close_matches(agent_filter, [a.lower() for a in all_agents], n=3, cutoff=0.5)
        if close:
            await update.message.reply_text(
                f"❓ No agent matching *{agent_filter}*. Did you mean:\n" +
                "\n".join(f"• {c}" for c in close),
                parse_mode=ParseMode.MARKDOWN,
            )
        else:
            await update.message.reply_text(f"❌ No agent matching *{agent_filter}*.", parse_mode=ParseMode.MARKDOWN)
        return

    # Find the actual agent name(s)
    agent_names = list(set(r["agent"] for r in matches))

    # Build per-client summary: total balance, overdue balance, SI count, card code
    by_client: dict[str, dict] = {}
    for r in matches:
        key = r["card_name"]
        if key not in by_client:
            by_client[key] = {
                "code": r.get("card_code", ""),
                "total": 0.0,
                "overdue": 0.0,
                "si_count": 0,
            }
        by_client[key]["total"] += r["balance"]
        by_client[key]["si_count"] += 1
        if r.get("days_due", 0) > 0:
            by_client[key]["overdue"] += r["balance"]

    total_outstanding = sum(c["total"] for c in by_client.values())
    total_sis = len(matches)
    agent_display = ', '.join(agent_names[:2]).upper()

    lines = []
    lines.append(f"👤 Agent: {agent_display}")
    lines.append(f"Clients: {len(by_client)} | SIs: {total_sis}")
    lines.append(f"Total Outstanding: {fmt_peso(total_outstanding)}")
    lines.append("")
    lines.append("Client Breakdown:")

    sorted_clients = sorted(by_client.items(), key=lambda x: x[1]["total"], reverse=True)
    for name, info in sorted_clients[:25]:
        code_prefix = f"<code>{info['code']}</code> " if info['code'] else ""
        lines.append(f"\u2022 {code_prefix}{name}")
        overdue_str = f" | 🔴 Overdue: {fmt_peso(info['overdue'])}" if info['overdue'] > 0 else ""
        lines.append(f"  Total: {fmt_peso(info['total'])}{overdue_str}")

    if len(by_client) > 25:
        lines.append(f"(Showing 25 of {len(by_client)} clients)")

    lines.append(f"\n{ar_source_footer()}")
    # Use plain text (no ParseMode) to avoid Markdown crashes from special chars in client names
    full_text = "\n".join(lines)
    # Chunk if needed
    MAX_LEN = 4000
    if len(full_text) <= MAX_LEN:
        await update.message.reply_text(full_text, parse_mode=ParseMode.HTML)
    else:
        chunks = []
        current = ""
        for line in lines:
            if len(current) + len(line) + 1 > MAX_LEN:
                chunks.append(current.rstrip())
                current = line + "\n"
            else:
                current += line + "\n"
        if current.strip():
            chunks.append(current.rstrip())
        for chunk in chunks:
            await update.message.reply_text(chunk, parse_mode=ParseMode.HTML)


async def cmd_arsearch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """Search clients by name in AR."""
    keyword = " ".join(context.args).strip() if context.args else ""
    if not keyword:
        await update.message.reply_text("Usage: `/arsearch <keyword>`\nExample: `/arsearch jollibee`", parse_mode=ParseMode.MARKDOWN)
        return

    if not store.ar_rows:
        await update.message.reply_text("⚠️ AR data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    q = keyword.lower()
    matches = [r for r in store.ar_rows if q in r["card_name"].lower() or q in r["card_code"].lower()]
    if not matches:
        await update.message.reply_text(f"❌ No clients matching *{keyword}*.", parse_mode=ParseMode.MARKDOWN)
        return

    by_client: dict[str, dict] = {}
    for r in matches:
        k = r["card_name"]
        if k not in by_client:
            by_client[k] = {"name": k, "code": r["card_code"], "total": 0.0, "agent": r["agent"]}
        by_client[k]["total"] += r["balance"]

    lines = [f"🔎 AR Search: {keyword} — {len(by_client)} client(s)\n"]
    for c in sorted(by_client.values(), key=lambda x: x["total"], reverse=True)[:20]:
        lines.append(f"\u2022 <code>{c['code']}</code> {c['name']}")
        lines.append(f"  Balance: {fmt_peso(c['total'])} | Agent: {c['agent'] or '—'}")

    if len(by_client) > 20:
        lines.append(f"Showing 20 of {len(by_client)} clients.")

    lines.append(f"\n{ar_source_footer()}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


async def cmd_arrefresh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """Force-refresh AR data only."""
    msg = await update.message.reply_text("🔄 Refreshing AR data from Google Drive...")
    try:
        refresh_ar_only()
        total = sum(r["balance"] for r in store.ar_rows)
        clients = len(set(r["card_name"] for r in store.ar_rows))
        ts = store.last_refresh.strftime("%m/%d/%Y %I:%M %p") if store.last_refresh else "—"
        await msg.edit_text(
            f"✅ AR data refreshed!\n"
            f"Clients: {clients} | SIs: {len(store.ar_rows)}\n"
            f"Total Outstanding: {fmt_peso(total)}\n\n"
            f"🕐 _Refreshed at {ts} PHT_",
            parse_mode=ParseMode.MARKDOWN,
        )
    except Exception as e:
        await msg.edit_text(f"❌ AR refresh failed: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Command handlers — AP
# ──────────────────────────────────────────────────────────────────────────────
async def cmd_ap_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """AP summary with aging buckets + top vendors."""
    if not store.ap_rows:
        await update.message.reply_text("⚠️ AP data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    total = sum(r["doc_total"] for r in store.ap_rows)
    vendors = len(set(r["card_name"] for r in store.ap_rows))
    txn_count = len(store.ap_rows)
    today = datetime.now(PHT).date()

    # AP aging by due date
    buckets = {"Current": 0.0, "1-30d": 0.0, "31-60d": 0.0, "61-90d": 0.0, "90+d": 0.0}
    for r in store.ap_rows:
        days_od = get_ap_days_overdue(r)
        if days_od <= 0:
            buckets["Current"] += r["doc_total"]
        elif days_od <= 30:
            buckets["1-30d"] += r["doc_total"]
        elif days_od <= 60:
            buckets["31-60d"] += r["doc_total"]
        elif days_od <= 90:
            buckets["61-90d"] += r["doc_total"]
        else:
            buckets["90+d"] += r["doc_total"]

    # Top vendors by amount
    by_vendor: dict[str, float] = {}
    for r in store.ap_rows:
        by_vendor[r["card_name"]] = by_vendor.get(r["card_name"], 0.0) + r["doc_total"]
    top_vendors = sorted(by_vendor.items(), key=lambda x: x[1], reverse=True)[:5]

    ts = store.last_refresh.strftime("%m/%d/%Y") if store.last_refresh else "—"

    lines = [
        "💳 *AP Summary (Unreleased Payments)*",
        f"_As of {ts} PHT_",
        "",
        f"💰 Total: *{fmt_peso(total)}*",
        f"🏢 Vendors: {vendors} | 📄 Transactions: {txn_count}",
        "",
        "*Aging Breakdown:*",
    ]
    for bucket, amount in buckets.items():
        pct = (amount / total * 100) if total > 0 else 0
        lines.append(f"  {bucket}: {fmt_peso(amount)} ({pct:.1f}%)")

    lines.append("")
    lines.append("*Top 5 Vendors:*")
    for vendor, amount in top_vendors:
        lines.append(f"  • {vendor}: {fmt_peso(amount)}")

    lines.append("")
    lines.append(ap_source_footer())
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


# /apsummary is the primary name; /ap_summary is the alias
async def cmd_apsummary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    await cmd_ap_summary(update, context)


async def cmd_ap(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    query = " ".join(context.args).strip() if context.args else ""
    if not query:
        await update.message.reply_text("Usage: `/ap <vendor name>`\nExample: `/ap san miguel`", parse_mode=ParseMode.MARKDOWN)
        return
    if not store.ap_rows:
        await update.message.reply_text("⚠️ AP data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    q = query.lower()
    matches = [r for r in store.ap_rows if q in r["card_name"].lower()]
    if not matches:
        all_vendors = list(set(r["card_name"] for r in store.ap_rows))
        close = difflib.get_close_matches(query, all_vendors, n=3, cutoff=0.5)
        if close:
            await update.message.reply_text(
                f"❓ No exact match for *{query}*. Did you mean:\n" +
                "\n".join(f"• {c}" for c in close),
                parse_mode=ParseMode.MARKDOWN,
            )
        else:
            await update.message.reply_text(f"❌ No AP records for *{query}*.", parse_mode=ParseMode.MARKDOWN)
        return

    by_vendor: dict[str, dict] = {}
    for r in matches:
        k = r["card_name"]
        if k not in by_vendor:
            by_vendor[k] = {"name": k, "total": 0.0, "txns": []}
        by_vendor[k]["total"] += r["doc_total"]
        by_vendor[k]["txns"].append(r)

    lines = [f"💳 AP: {query} — {len(by_vendor)} vendor(s)\n"]
    for vendor in sorted(by_vendor.values(), key=lambda x: x["total"], reverse=True)[:5]:
        lines.append(f"🏢 {vendor['name']}")
        lines.append(f"   Total Unreleased: {fmt_peso(vendor['total'])}")
        for txn in vendor["txns"][:10]:
            due = fmt_date(txn["due_date"])
            days_od = get_ap_days_overdue(txn)
            od_tag = f" ⚠️ {days_od}d overdue" if days_od > 0 else ""
            lines.append(f"   • Doc {txn['doc_num']} | Due: {due} | {fmt_peso(txn['doc_total'])}{od_tag}")
            if txn["comments"]:
                lines.append(f"     {txn['comments'][:80]}")
        if len(vendor["txns"]) > 10:
            lines.append(f"   (+{len(vendor['txns'])-10} more transactions)")
        lines.append("")

    lines.append(ap_source_footer())
    full_text = "\n".join(lines)
    MAX_LEN = 4000
    if len(full_text) <= MAX_LEN:
        await update.message.reply_text(full_text)
    else:
        chunks, current = [], ""
        for line in lines:
            if len(current) + len(line) + 1 > MAX_LEN:
                chunks.append(current.rstrip())
                current = line + "\n"
            else:
                current += line + "\n"
        if current.strip():
            chunks.append(current.rstrip())
        for chunk in chunks:
            await update.message.reply_text(chunk)


# /vendor is an alias for /ap
async def cmd_vendor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    await cmd_ap(update, context)


async def cmd_apaging(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """AP aging by bucket."""
    if not store.ap_rows:
        await update.message.reply_text("⚠️ AP data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    buckets = {"Current (not due)": 0.0, "1-30d overdue": 0.0, "31-60d overdue": 0.0, "61-90d overdue": 0.0, "90+d overdue": 0.0}
    bucket_counts = {k: 0 for k in buckets}

    for r in store.ap_rows:
        days_od = get_ap_days_overdue(r)
        if days_od <= 0:
            buckets["Current (not due)"] += r["doc_total"]
            bucket_counts["Current (not due)"] += 1
        elif days_od <= 30:
            buckets["1-30d overdue"] += r["doc_total"]
            bucket_counts["1-30d overdue"] += 1
        elif days_od <= 60:
            buckets["31-60d overdue"] += r["doc_total"]
            bucket_counts["31-60d overdue"] += 1
        elif days_od <= 90:
            buckets["61-90d overdue"] += r["doc_total"]
            bucket_counts["61-90d overdue"] += 1
        else:
            buckets["90+d overdue"] += r["doc_total"]
            bucket_counts["90+d overdue"] += 1

    total = sum(buckets.values())
    lines = ["💳 *AP Aging (Unreleased Payments)*\n"]
    for bucket, amount in buckets.items():
        pct = (amount / total * 100) if total > 0 else 0
        lines.append(f"• *{bucket}*: {fmt_peso(amount)} ({pct:.1f}%) — {bucket_counts[bucket]} docs")

    lines.append(f"\n💰 *Total Unreleased: {fmt_peso(total)}*")
    lines.append(ap_source_footer())
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


async def cmd_apoverdue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """Vendors with payments overdue 61+ days."""
    if not store.ap_rows:
        await update.message.reply_text("⚠️ AP data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    overdue = [r for r in store.ap_rows if get_ap_days_overdue(r) >= 61]
    if not overdue:
        await update.message.reply_text("✅ No AP payments overdue 61+ days.", parse_mode=ParseMode.MARKDOWN)
        return

    by_vendor: dict[str, dict] = {}
    for r in overdue:
        k = r["card_name"]
        if k not in by_vendor:
            by_vendor[k] = {"name": k, "total": 0.0, "max_days": 0, "count": 0}
        by_vendor[k]["total"] += r["doc_total"]
        days_od = get_ap_days_overdue(r)
        by_vendor[k]["max_days"] = max(by_vendor[k]["max_days"], days_od)
        by_vendor[k]["count"] += 1

    vendors = sorted(by_vendor.values(), key=lambda x: x["total"], reverse=True)
    total_overdue = sum(r["doc_total"] for r in overdue)

    lines = [f"🚨 AP Overdue 61+ Days ({len(vendors)} vendors)\n"]
    for i, v in enumerate(vendors[:15], 1):
        lines.append(f"{i}. {v['name']}")
        lines.append(f"   {fmt_peso(v['total'])} | Max: {v['max_days']}d | {v['count']} docs")

    if len(vendors) > 15:
        lines.append(f"\nShowing 15 of {len(vendors)} vendors.")

    lines.append(f"\n💰 Total Overdue: {fmt_peso(total_overdue)}")
    lines.append(ap_source_footer())
    await update.message.reply_text("\n".join(lines))


async def cmd_aptop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """Top N vendors by total unreleased amount (default 10)."""
    n = 10
    if context.args:
        try:
            n = int(context.args[0])
            n = max(1, min(n, 50))
        except ValueError:
            await update.message.reply_text("Usage: `/aptop [n]`\nExample: `/aptop 20`", parse_mode=ParseMode.MARKDOWN)
            return

    if not store.ap_rows:
        await update.message.reply_text("⚠️ AP data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    by_vendor: dict[str, float] = {}
    for r in store.ap_rows:
        by_vendor[r["card_name"]] = by_vendor.get(r["card_name"], 0.0) + r["doc_total"]

    top = sorted(by_vendor.items(), key=lambda x: x[1], reverse=True)[:n]
    grand_total = sum(by_vendor.values())

    lines = [f"🏆 Top {n} Vendors by AP Amount\n"]
    for i, (name, amount) in enumerate(top, 1):
        pct = (amount / grand_total * 100) if grand_total > 0 else 0
        lines.append(f"{i}. {name}")
        lines.append(f"   {fmt_peso(amount)} ({pct:.1f}%)")

    lines.append(f"\n💰 Grand Total: {fmt_peso(grand_total)}")
    lines.append(ap_source_footer())
    await update.message.reply_text("\n".join(lines))


async def cmd_due_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    if not store.ap_rows:
        await update.message.reply_text("⚠️ AP data not loaded yet. Try `/refresh`.", parse_mode=ParseMode.MARKDOWN)
        return

    today = datetime.now(PHT).date()
    week_end = today + timedelta(days=7)

    due_this_week = []
    for r in store.ap_rows:
        if r["due_date"] is None:
            continue
        d = r["due_date"]
        if isinstance(d, datetime):
            d = d.date()
        if isinstance(d, date) and today <= d <= week_end:
            due_this_week.append({**r, "_due_date_obj": d})

    if not due_this_week:
        await update.message.reply_text(
            f"✅ No AP payments due in the next 7 days (through {week_end.strftime('%m/%d/%Y')}).",
            parse_mode=ParseMode.MARKDOWN,
        )
        return

    due_this_week.sort(key=lambda x: x["_due_date_obj"])
    total = sum(r["doc_total"] for r in due_this_week)

    lines = [f"📅 AP Due This Week ({today.strftime('%m/%d')} – {week_end.strftime('%m/%d')})\n"]
    for r in due_this_week[:20]:
        lines.append(f"• {r['card_name']}")
        lines.append(f"  Doc {r['doc_num']} | Due: {fmt_date(r['due_date'])} | {fmt_peso(r['doc_total'])}")

    if len(due_this_week) > 20:
        lines.append(f"\nShowing 20 of {len(due_this_week)} payments.")

    lines.append(f"\n💰 Total Due: {fmt_peso(total)}")
    lines.append(ap_source_footer())
    await update.message.reply_text("\n".join(lines))


async def cmd_aprefresh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """Force-refresh AP data only."""
    msg = await update.message.reply_text("🔄 Refreshing AP data from Google Drive...")
    try:
        refresh_ap_only()
        total = sum(r["doc_total"] for r in store.ap_rows)
        vendors = len(set(r["card_name"] for r in store.ap_rows))
        ts = store.last_refresh.strftime("%m/%d/%Y %I:%M %p") if store.last_refresh else "—"
        src_ts = store.ap_source_ts or ts
        await msg.edit_text(
            f"✅ AP data refreshed!\n"
            f"Vendors: {vendors} | Transactions: {len(store.ap_rows)}\n"
            f"Total Unreleased: {fmt_peso(total)}\n\n"
            f"🕐 _Source data as of {src_ts} PHT_",
            parse_mode=ParseMode.MARKDOWN,
        )
    except Exception as e:
        await msg.edit_text(f"❌ AP refresh failed: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Sales Performance (Portal DB)
# ──────────────────────────────────────────────────────────────────────────────

def _get_portal_conn():
    """Open a read-only connection to the Ham Portal MySQL/TiDB database."""
    if not PYMYSQL_AVAILABLE:
        raise RuntimeError("pymysql not installed")
    if not PORTAL_DB_URL:
        raise RuntimeError("PORTAL_DB_URL not configured")
    # Parse mysql://user:pass@host:port/db?ssl={...}
    import urllib.parse
    parsed = urllib.parse.urlparse(PORTAL_DB_URL)
    host = parsed.hostname
    port = parsed.port or 4000
    user = parsed.username
    password = parsed.password
    db = parsed.path.lstrip("/")
    ssl_arg = None
    qs = urllib.parse.parse_qs(parsed.query)
    if "ssl" in qs:
        try:
            ssl_val = json.loads(qs["ssl"][0])
            ssl_arg = ssl_val
        except Exception:
            ssl_arg = {"rejectUnauthorized": True}
    conn = pymysql.connect(
        host=host, port=port, user=user, password=password, database=db,
        ssl=ssl_arg, cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=10, read_timeout=15,
    )
    return conn


MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _current_fy_year() -> int:
    """Fiscal year: Jul–Jun. FY2027 = Jul 2026 – Jun 2027."""
    now = datetime.now(PHT)
    return now.year + 1 if now.month >= 7 else now.year


def _fy_month_range(fy_year: int) -> tuple[tuple[int, int], tuple[int, int]]:
    """Return (start_year, start_month), (end_year, end_month) for a fiscal year (Jul–Jun)."""
    return (fy_year - 1, 7), (fy_year, 6)


def _parse_sales_args(args: list[str], now) -> tuple[str, int | None, int | None, str]:
    """Parse /sales args into (area_filter, month_filter, year_filter, cat_filter)."""
    cat_filter = ""
    remaining_args = []
    for a in args:
        if a.lower().startswith("cat:"):
            cat_filter = a[4:].upper().strip()
        else:
            remaining_args.append(a)
    area_tokens = []
    month_filter = None
    year_filter = None
    for token in remaining_args:
        try:
            n = int(token)
            if 1 <= n <= 12:
                month_filter = n
                year_filter = now.year
                continue
        except ValueError:
            pass
        mn = token.strip().lower()[:3].capitalize()
        if mn in MONTH_NAMES:
            month_filter = MONTH_NAMES.index(mn) + 1
            year_filter = now.year
            continue
        area_tokens.append(token)
    return " ".join(area_tokens).strip(), month_filter, year_filter, cat_filter


async def _do_sales_query(
    reply_target,  # message or callback query
    area_exact: str,  # exact area name (already resolved)
    month_filter: int | None,
    year_filter: int | None,
    cat_filter: str,
    edit: bool = False,
):
    """Run the sales query for a specific (exact) area and render the result."""
    try:
        conn = _get_portal_conn()
        with conn:
            with conn.cursor() as cur:
                params: list = []
                where_clauses: list[str] = []
                if area_exact:
                    where_clauses.append("location = %s")
                    params.append(area_exact)
                if month_filter and year_filter:
                    where_clauses.append("periodYear = %s AND periodMonth = %s")
                    params.extend([year_filter, month_filter])
                if cat_filter:
                    where_clauses.append("cat = %s")
                    params.append(cat_filter)
                where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
                cur.execute(
                    f"""
                    SELECT location AS segment,
                           COALESCE(SUM(CAST(rowTotal AS DECIMAL(16,2))), 0) AS gross_revenue,
                           COUNT(*) AS tx_count
                    FROM sales_transactions
                    {where_sql}
                    GROUP BY location
                    ORDER BY gross_revenue DESC
                    """,
                    params
                )
                rows = cur.fetchall()

                # Fetch CM deductions per segment for the same period
                cm_params: list = []
                cm_where: list[str] = []
                if area_exact:
                    cm_where.append("segment = %s")
                    cm_params.append(area_exact)
                if month_filter and year_filter:
                    cm_where.append("periodYear = %s AND periodMonth = %s")
                    cm_params.extend([year_filter, month_filter])
                cm_where_sql = ("WHERE " + " AND ".join(cm_where)) if cm_where else ""
                cur.execute(
                    f"""
                    SELECT segment,
                           COALESCE(SUM(CAST(lineTotal AS DECIMAL(16,2))), 0) AS cm_total
                    FROM credit_memo_lines
                    {cm_where_sql}
                    GROUP BY segment
                    """,
                    cm_params
                )
                cm_rows = cur.fetchall()
                cm_map = {r["segment"]: float(r["cm_total"]) for r in cm_rows if r["segment"]}
                if month_filter and year_filter:
                    cur.execute(
                        """
                        SELECT segment,
                               COALESCE(SUM(CAST(targetAmount AS DECIMAL(16,2))), 0) AS target
                        FROM financial_targets
                        WHERE periodYear = %s AND periodMonth = %s
                        GROUP BY segment
                        """,
                        (year_filter, month_filter)
                    )
                else:
                    cur.execute(
                        """
                        SELECT segment,
                               COALESCE(SUM(CAST(targetAmount AS DECIMAL(16,2))), 0) AS target
                        FROM financial_targets
                        GROUP BY segment
                        """
                    )
                target_rows = cur.fetchall()
        target_map = {r["segment"]: float(r["target"]) for r in target_rows}
        total_gross = sum(float(r["gross_revenue"]) for r in rows)
        total_cm = sum(cm_map.values())
        total_net = total_gross - total_cm
        period_label = f"{MONTH_NAMES[month_filter-1]} {year_filter}" if month_filter else "All months"
        area_label = f" | {area_exact}" if area_exact else ""
        cat_label = f" | Cat: {cat_filter}" if cat_filter else ""
        lines = [
            f"📊 *Net Sales by Area*",
            f"_{period_label}{area_label}{cat_label}_",
            "",
        ]
        if not rows:
            lines.append("_No sales data found._")
        else:
            for r in rows:
                seg = r["segment"] or "Unknown"
                gross = float(r["gross_revenue"])
                cm_deduct = cm_map.get(seg, 0.0)
                net = gross - cm_deduct
                target = target_map.get(seg, 0)
                pct = f" ({net/target*100:.0f}% of target)" if target > 0 else ""
                cm_note = f" _(CM: -{fmt_peso(cm_deduct)})_" if cm_deduct > 0 else ""
                lines.append(f"• *{seg}*: {fmt_peso(net)}{pct}{cm_note}")
            lines.append("")
            lines.append(f"💰 *Net Total: {fmt_peso(total_net)}*")
            if total_cm > 0:
                lines.append(f"_Gross: {fmt_peso(total_gross)} | CM deducted: -{fmt_peso(total_cm)}_")
        lines.append("")
        lines.append("_Source: Ham Portal DB (HANA + B1 consolidated)_")
        text = "\n".join(lines)
        if edit:
            await reply_target.edit_message_text(text, parse_mode=ParseMode.MARKDOWN)
        else:
            await reply_target.edit_text(text, parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        logger.error(f"_do_sales_query error: {e}")
        err_text = f"❌ Sales query failed: {e}"
        if edit:
            await reply_target.edit_message_text(err_text)
        else:
            await reply_target.edit_text(err_text)


async def cmd_sales(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """/sales [area] [month] [cat:CATEGORY] — Sales by area with optional month filter.
    Examples:
      /sales                    → all areas, all months
      /sales cebu               → Cebu only (exact match, or picker if ambiguous)
      /sales cebu july          → Cebu, July (current year)
      /sales july               → all areas, July
      /sales 7                  → all areas, month 7
    Optional: cat:PROCESSED | cat:TRADING | cat:SEASONAL | cat:DAIRY | cat:BUNDLE | cat:OTHERS
    """
    now = datetime.now(PHT)
    args = context.args or []
    area_filter, month_filter, year_filter, cat_filter = _parse_sales_args(args, now)

    msg = await update.message.reply_text("📊 Fetching sales data...")

    # No area filter — show all areas directly
    if not area_filter:
        await _do_sales_query(msg, "", month_filter, year_filter, cat_filter)
        return

    try:
        conn = _get_portal_conn()
        with conn:
            with conn.cursor() as cur:
                # Find all matching area names (exact first, then partial)
                cur.execute(
                    "SELECT DISTINCT location FROM sales_transactions "
                    "WHERE LOWER(location) = LOWER(%s) ORDER BY location",
                    (area_filter,)
                )
                exact_matches = [r["location"] for r in cur.fetchall()]
                if exact_matches:
                    matched_areas = exact_matches
                else:
                    cur.execute(
                        "SELECT DISTINCT location FROM sales_transactions "
                        "WHERE location LIKE %s ORDER BY location",
                        (f"%{area_filter}%",)
                    )
                    matched_areas = [r["location"] for r in cur.fetchall()]
    except Exception as e:
        logger.error(f"cmd_sales area lookup error: {e}")
        await msg.edit_text(f"❌ Sales query failed: {e}")
        return

    if not matched_areas:
        await msg.edit_text(f"❌ No area found matching *{area_filter}*. Try `/sales` to see all areas.", parse_mode=ParseMode.MARKDOWN)
        return

    if len(matched_areas) == 1:
        # Exactly one match — run query directly
        await _do_sales_query(msg, matched_areas[0], month_filter, year_filter, cat_filter)
        return

    # Multiple matches — show inline keyboard picker
    month_part = f"|{month_filter}|{year_filter}" if month_filter else "||"
    cat_part = f"|{cat_filter}" if cat_filter else "|"
    buttons = [
        [InlineKeyboardButton(area, callback_data=f"sales_area|{area}{month_part}{cat_part}")]
        for area in matched_areas
    ]
    keyboard = InlineKeyboardMarkup(buttons)
    period_label = f"{MONTH_NAMES[month_filter-1]} {year_filter}" if month_filter else "all months"
    await msg.edit_text(
        f"📊 Found *{len(matched_areas)} areas* matching _{area_filter}_. Which one?"
        f"\n_Period: {period_label}_",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=keyboard,
    )


async def cb_sales_area_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback: user tapped an area button from /sales disambiguation picker."""
    query = update.callback_query
    user = query.from_user
    chat = query.message.chat if query.message is not None else None
    if chat is not None:
        log_chat(chat.id, chat.type, chat.title)
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if chat is not None and is_whitelisted_group(chat.id):
                pass  # whitelisted group: fully open for all members
            elif ACCESS_MODE == "hard":
                await _send_blocked_notice(query, str(user.id))
                return
            elif ACCESS_MODE == "soft":
                await _send_blocked_notice(query, str(user.id))
                # soft: continue
    await query.answer()
    # Format: sales_area|<area>|<month_or_empty>|<year_or_empty>|<cat_or_empty>
    parts = query.data.split("|", 4)
    if len(parts) < 5:
        await query.edit_message_text("❌ Invalid selection.")
        return
    _, area_exact, month_str, year_str, cat_filter = parts
    month_filter = int(month_str) if month_str else None
    year_filter = int(year_str) if year_str else None
    await _do_sales_query(query, area_exact, month_filter, year_filter, cat_filter, edit=True)


async def cmd_salesagent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """/salesagent [name] — Sales by agent from portal DB."""
    args = context.args or []
    agent_filter = " ".join(args).strip() if args else ""

    msg = await update.message.reply_text("👤 Fetching agent sales...")
    try:
        conn = _get_portal_conn()
        with conn:
            with conn.cursor() as cur:
                if agent_filter:
                    cur.execute(
                        """
                        SELECT agentName,
                               COALESCE(SUM(CAST(rowTotal AS DECIMAL(16,2))), 0) AS revenue,
                               COUNT(*) AS tx_count
                        FROM sales_transactions
                        WHERE agentName LIKE %s
                        GROUP BY agentName
                        ORDER BY revenue DESC
                        """,
                        (f"%{agent_filter}%",)
                    )
                else:
                    cur.execute(
                        """
                        SELECT agentName,
                               COALESCE(SUM(CAST(rowTotal AS DECIMAL(16,2))), 0) AS revenue,
                               COUNT(*) AS tx_count
                        FROM sales_transactions
                        GROUP BY agentName
                        ORDER BY revenue DESC
                        LIMIT 20
                        """
                    )
                rows = cur.fetchall()

        total_rev = sum(float(r["revenue"]) for r in rows)
        lines = [
            f"👤 *Sales by Agent*",
            f"_All available data_",
            "",
        ]
        if not rows:
            lines.append("_No agent data found._")
        else:
            for i, r in enumerate(rows, 1):
                name = r["agentName"] or "Unknown"
                rev = float(r["revenue"])
                pct = f" ({rev/total_rev*100:.1f}%" + ")" if total_rev > 0 else ""
                lines.append(f"{i}. *{name}*: {fmt_peso(rev)}{pct}")
            lines.append("")
            lines.append(f"💰 *Total: {fmt_peso(total_rev)}*")
        lines.append("")
        lines.append("_Source: Ham Portal DB (HANA + B1 consolidated)_")
        await msg.edit_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        logger.error(f"cmd_salesagent error: {e}")
        await msg.edit_text(f"❌ Agent sales query failed: {e}")


async def cmd_salestarget(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """/salestarget [month] — Segment performance vs target."""
    now = datetime.now(PHT)

    # Optional: filter to a specific month
    month_filter = None
    year_filter = None
    if context.args:
        try:
            month_filter = int(context.args[0])
            year_filter = int(context.args[1]) if len(context.args) > 1 else now.year
        except (ValueError, IndexError):
            # Try month name
            mn = context.args[0].strip().lower()[:3].capitalize()
            if mn in MONTH_NAMES:
                month_filter = MONTH_NAMES.index(mn) + 1
                year_filter = now.year

    msg = await update.message.reply_text("🎯 Fetching targets vs actuals...")
    try:
        conn = _get_portal_conn()
        with conn:
            with conn.cursor() as cur:
                if month_filter and year_filter:
                    cur.execute(
                        """
                        SELECT location AS segment,
                               COALESCE(SUM(CAST(rowTotal AS DECIMAL(16,2))), 0) AS revenue
                        FROM sales_transactions
                        WHERE periodYear = %s AND periodMonth = %s
                        GROUP BY location
                        """,
                        (year_filter, month_filter)
                    )
                    rev_rows = cur.fetchall()
                    cur.execute(
                        """
                        SELECT segment,
                               COALESCE(SUM(CAST(targetAmount AS DECIMAL(16,2))), 0) AS target
                        FROM financial_targets
                        WHERE periodYear = %s AND periodMonth = %s
                        GROUP BY segment
                        """,
                        (year_filter, month_filter)
                    )
                    tgt_rows = cur.fetchall()
                    period_label = f"{MONTH_NAMES[month_filter-1]} {year_filter}"
                else:
                    cur.execute(
                        """
                        SELECT location AS segment,
                               COALESCE(SUM(CAST(rowTotal AS DECIMAL(16,2))), 0) AS revenue
                        FROM sales_transactions
                        GROUP BY location
                        """
                    )
                    rev_rows = cur.fetchall()
                    cur.execute(
                        """
                        SELECT segment,
                               COALESCE(SUM(CAST(targetAmount AS DECIMAL(16,2))), 0) AS target
                        FROM financial_targets
                        GROUP BY segment
                        """
                    )
                    tgt_rows = cur.fetchall()
                    period_label = "All available data"

        rev_map = {r["segment"]: float(r["revenue"]) for r in rev_rows}
        tgt_map = {r["segment"]: float(r["target"]) for r in tgt_rows}
        all_segs = sorted(set(list(rev_map.keys()) + list(tgt_map.keys())))

        total_rev = sum(rev_map.values())
        total_tgt = sum(tgt_map.values())

        lines = [
            f"🎯 *Sales vs Target — {period_label}*",
            "",
        ]
        if not all_segs:
            lines.append("_No data found._")
        else:
            for seg in all_segs:
                rev = rev_map.get(seg, 0)
                tgt = tgt_map.get(seg, 0)
                if tgt > 0:
                    pct = rev / tgt * 100
                    bar = "✅" if pct >= 100 else ("🟡" if pct >= 75 else "🔴")
                    lines.append(f"{bar} *{seg}*")
                    lines.append(f"   Actual: {fmt_peso(rev)} / Target: {fmt_peso(tgt)} ({pct:.0f}%)")
                else:
                    lines.append(f"📌 *{seg}*: {fmt_peso(rev)} _(no target set)_")
            lines.append("")
            if total_tgt > 0:
                overall_pct = total_rev / total_tgt * 100
                lines.append(f"💰 *Total: {fmt_peso(total_rev)} / {fmt_peso(total_tgt)} ({overall_pct:.0f}%)*")
            else:
                lines.append(f"💰 *Total: {fmt_peso(total_rev)}*")
        lines.append("")
        lines.append("_Source: Ham Portal DB (HANA + B1 consolidated)_")
        await msg.edit_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        logger.error(f"cmd_salestarget error: {e}")
        await msg.edit_text(f"❌ Target query failed: {e}")


async def cmd_salesproduct(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """/salesproduct [keyword] [cat:CATEGORY] — Top products by revenue from portal DB.
    Optional: cat:PROCESSED | cat:TRADING | cat:SEASONAL | cat:DAIRY | cat:BUNDLE | cat:OTHERS
    """
    args = context.args or []
    cat_filter = ""
    remaining_args = []
    for a in args:
        if a.lower().startswith("cat:"):
            cat_filter = a[4:].upper().strip()
        else:
            remaining_args.append(a)
    kw = " ".join(remaining_args).strip()
    msg = await update.message.reply_text("📦 Fetching product sales...")
    try:
        conn = _get_portal_conn()
        with conn:
            with conn.cursor() as cur:
                params = []
                where_clauses = []
                if kw:
                    where_clauses.append("(itemDescription LIKE %s OR itemCategory LIKE %s)")
                    params.extend([f"%{kw}%", f"%{kw}%"])
                if cat_filter:
                    where_clauses.append("cat = %s")
                    params.append(cat_filter)
                where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
                cur.execute(
                    f"""
                    SELECT itemDescription AS product,
                           cat AS cat_label,
                           COALESCE(SUM(CAST(rowTotal AS DECIMAL(16,2))), 0) AS revenue,
                           COALESCE(SUM(CAST(quantity AS DECIMAL(12,3))), 0) AS qty
                    FROM sales_transactions
                    {where_sql}
                    GROUP BY itemDescription, cat
                    ORDER BY revenue DESC
                    LIMIT 20
                    """,
                    params
                )
                rows = cur.fetchall()

        total_rev = sum(float(r["revenue"]) for r in rows)
        cat_label = f" | {cat_filter}" if cat_filter else ""
        lines = [
            f"📦 *Top Products by Revenue{cat_label}*",
            f"_All available data_",
            "",
        ]
        if not rows:
            lines.append("_No product data found._")
        else:
            for i, r in enumerate(rows, 1):
                prod = r["product"] or "Unknown"
                cat = r["cat_label"] or ""
                rev = float(r["revenue"])
                qty = float(r["qty"])
                cat_str = f" _{cat}_" if cat else ""
                lines.append(f"{i}. *{prod}*{cat_str}")
                lines.append(f"   {fmt_peso(rev)} | {qty:,.1f} units")
            lines.append("")
            lines.append(f"💰 *Top 20 Total: {fmt_peso(total_rev)}*")
        lines.append("")
        lines.append("_Source: Ham Portal DB (HANA + B1 consolidated)_")
        await msg.edit_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        logger.error(f"cmd_salesproduct error: {e}")
        await msg.edit_text(f"❌ Product sales query failed: {e}")


async def cmd_salesmonth(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """/salesmonth — Monthly net revenue trend (current year, grouped by posting date)."""
    msg = await update.message.reply_text("📅 Fetching monthly trend...")
    try:
        conn = _get_portal_conn()
        with conn:
            with conn.cursor() as cur:
                # Gross sales grouped by posting date year+month
                cur.execute(
                    """
                    SELECT YEAR(postingDate) AS yr,
                           MONTH(postingDate) AS mo,
                           COALESCE(SUM(CAST(rowTotal AS DECIMAL(16,2))), 0) AS gross_revenue,
                           COUNT(*) AS tx_count
                    FROM sales_transactions
                    WHERE YEAR(postingDate) = YEAR(CURDATE())
                    GROUP BY YEAR(postingDate), MONTH(postingDate)
                    ORDER BY YEAR(postingDate), MONTH(postingDate)
                    """
                )
                rows = cur.fetchall()

                # CM deductions grouped by posting date year+month
                cur.execute(
                    """
                    SELECT periodYear AS yr,
                           periodMonth AS mo,
                           COALESCE(SUM(CAST(lineTotal AS DECIMAL(16,2))), 0) AS cm_total
                    FROM credit_memo_lines
                    WHERE periodYear = YEAR(CURDATE())
                    GROUP BY periodYear, periodMonth
                    """
                )
                cm_rows = cur.fetchall()
                cm_map = {(r["yr"], r["mo"]): float(r["cm_total"]) for r in cm_rows}

        total_gross = sum(float(r["gross_revenue"]) for r in rows)
        total_cm = sum(cm_map.values())
        total_net = total_gross - total_cm
        cur_year = __import__('datetime').date.today().year
        lines = [
            f"📅 *Monthly Net Revenue — {cur_year}*",
            f"_Grouped by posting date | CM deducted_",
            "",
        ]
        if not rows:
            lines.append("_No monthly data found._")
        else:
            max_net = max(
                float(r["gross_revenue"]) - cm_map.get((r["yr"], r["mo"]), 0.0)
                for r in rows
            ) if rows else 1
            for r in rows:
                yr = r["yr"]
                mo = r["mo"]
                gross = float(r["gross_revenue"])
                cm_deduct = cm_map.get((yr, mo), 0.0)
                net = gross - cm_deduct
                label = f"{MONTH_NAMES[mo-1]} {yr}"
                bar_len = int(net / max_net * 10) if max_net > 0 else 0
                bar = "█" * bar_len
                cm_note = f" _(CM:-{fmt_peso(cm_deduct)})_" if cm_deduct > 0 else ""
                lines.append(f"`{label:<10}` {bar} {fmt_peso(net)}{cm_note}")
            lines.append("")
            lines.append(f"💰 *Net Total: {fmt_peso(total_net)}*")
            if total_cm > 0:
                lines.append(f"_Gross: {fmt_peso(total_gross)} | CM: -{fmt_peso(total_cm)}_")
        lines.append("")
        lines.append("_Source: Ham Portal DB (HANA + B1 consolidated)_")
        await msg.edit_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        logger.error(f"cmd_salesmonth error: {e}")
        await msg.edit_text(f"❌ Monthly sales query failed: {e}")


async def cmd_slowmoving(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _access_gate(update, context):
        return
    """/slowmoving [days] — Items with stock but no sales in N days (default 30)."""
    days = 30
    if context.args and context.args[0].isdigit():
        days = int(context.args[0])

    msg = await update.message.reply_text(f"🐢 Identifying slow-moving items (last {days} days)...")
    await _do_slowmoving(msg, days, show_all=False)


async def _do_slowmoving(update_or_query, days: int, show_all: bool = False):
    try:
        # 1. Get all items with current stock
        with store._lock:
            inventory = store.inventory
        
        if not inventory:
            text = "📭 Inventory is empty. Run /refresh first."
            if hasattr(update_or_query, 'edit_text'): await update_or_query.edit_text(text)
            else: await update_or_query.edit_message_text(text)
            return

        # Filter inventory to exclude on-hold, FA, Engineering, and SP items before grouping
        # Exclude any warehouse containing "FA" or "ENGINEERING" in name or code
        # Exclude item codes starting with "SP"
        filtered_inv = [
            r for r in inventory 
            if r["whs_name"] not in ON_HOLD_WAREHOUSES and 
               r.get("whs_code") != "WDR12A" and
               "FA" not in r["whs_name"].upper() and
               "FA" not in r.get("whs_code", "").upper() and
               "ENGINEERING" not in r["whs_name"].upper() and
               not str(r["item_no"]).upper().startswith("SP")
        ]
        by_item = group_inventory_by_item(filtered_inv)
        items_with_stock = {k: v for k, v in by_item.items() if v["total"] > 0}
        
        # 2. Get sales from Portal DB for the last N days
        # We try to get both itemCode and itemDescription for better matching
        conn = _get_portal_conn()
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT 
                        itemCode,
                        itemDescription AS product,
                        SUM(CAST(quantity AS DECIMAL(12,3))) AS total_sold
                    FROM sales_transactions
                    WHERE postingDate >= DATE_SUB(CURDATE(), INTERVAL %s DAY)
                    GROUP BY itemCode, itemDescription
                    """,
                    (days,)
                )
                sales_rows = cur.fetchall()
        
        # Build two maps for matching: by code and by cleaned description
        sales_by_code = {}
        sales_by_desc = {}
        for r in sales_rows:
            qty = float(r["total_sold"])
            if r["itemCode"]:
                code = str(r["itemCode"]).strip().upper()
                sales_by_code[code] = sales_by_code.get(code, 0.0) + qty
            if r["product"]:
                # Clean description: lowercase, remove extra spaces
                desc = " ".join(str(r["product"]).lower().split())
                sales_by_desc[desc] = sales_by_desc.get(desc, 0.0) + qty
        
        # 3. Identify slow movers
        slow_movers = []
        for item_no, data in items_with_stock.items():
            code = str(item_no).strip().upper()
            desc = " ".join(str(data["desc"]).lower().split())
            
            # Check by code first, then by description
            sold = sales_by_code.get(code, 0.0)
            if sold == 0:
                sold = sales_by_desc.get(desc, 0.0)
            
            if sold == 0:
                slow_movers.append({
                    "item_no": item_no,
                    "desc": data["desc"],
                    "stock": data["total"]
                })
        
        slow_movers.sort(key=lambda x: x["stock"], reverse=True)
        
        PREVIEW_LIMIT = 20
        lines = [
            f"🐢 *Slow-Moving Items (Last {days} Days)*",
            f"_Items with stock but ZERO sales recorded._",
            "",
        ]
        
        if not slow_movers:
            lines.append("✨ *No slow-moving items found!* All stocked items have recent sales.")
        else:
            display_items = slow_movers if show_all else slow_movers[:PREVIEW_LIMIT]
            for i, item in enumerate(display_items, 1):
                lines.append(f"{i}. *{item['desc']}* (`{item['item_no']}`): {item['stock']:,.1f}")
            
            if not show_all and len(slow_movers) > PREVIEW_LIMIT:
                lines.append(f"\n_Showing {PREVIEW_LIMIT} of {len(slow_movers)} items_")
            
            lines.append("")
            lines.append(f"📦 *Total Slow Items: {len(slow_movers)}*")
            
        lines.append("")
        lines.append(inv_source_footer())
        
        text = "\n".join(lines)
        reply_markup = None
        if not show_all and len(slow_movers) > PREVIEW_LIMIT:
            cb_data = f"slow_all|{days}"
            reply_markup = InlineKeyboardMarkup([[InlineKeyboardButton(
                f"📋 Show All ({len(slow_movers)} items)", callback_data=cb_data
            )]])

        if show_all:
            # Chunking logic for long lists
            MAX_LEN = 3800
            chunks, cur = [], ""
            for line in lines:
                addition = ("\n" + line) if cur else line
                if len(cur) + len(addition) > MAX_LEN:
                    chunks.append(cur)
                    cur = line
                else:
                    cur += addition
            if cur: chunks.append(cur)
            
            first = True
            for chunk in chunks:
                if first:
                    if hasattr(update_or_query, 'edit_text'): await update_or_query.edit_text(chunk, parse_mode=ParseMode.MARKDOWN)
                    else: await update_or_query.edit_message_text(chunk, parse_mode=ParseMode.MARKDOWN)
                else:
                    await update_or_query.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
                first = False
        else:
            if hasattr(update_or_query, 'edit_text'): await update_or_query.edit_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
            else: await update_or_query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)

    except Exception as e:
        logger.error(f"_do_slowmoving error: {e}")
        err_text = f"❌ Failed to calculate slow-moving items: {e}"
        if hasattr(update_or_query, 'edit_text'): await update_or_query.edit_text(err_text)
        else: await update_or_query.edit_message_text(err_text)


async def cb_slowmoving_show_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback: show all slow-moving items."""
    query = update.callback_query
    user = query.from_user
    chat = query.message.chat if query.message is not None else None
    if chat is not None: log_chat(chat.id, chat.type, chat.title)
    if user is not None:
        log_user(user.id, user.username, user.full_name)
        if not is_registered(user.id):
            if chat is not None and is_whitelisted_group(chat.id): pass
            elif ACCESS_MODE == "hard":
                await _send_blocked_notice(query, str(user.id))
                return
            elif ACCESS_MODE == "soft":
                await _send_blocked_notice(query, str(user.id))
    await query.answer()
    days = int(query.data.split("|")[1]) if "|" in query.data else 30
    await _do_slowmoving(query, days, show_all=True)


# ── Free-text handler ─────────────────────────────────────────────────────────
async def handle_free_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip() if update.message.text else ""
    if not text:
        return

    # Ignore messages prefixed with // — treat as private notes
    if text.startswith("//"):
        return

    if text.startswith("/"):
        user = update.effective_user
        if user is not None:
            log_user(user.id, user.username, user.full_name)
            if not is_registered(user.id) and ACCESS_MODE == "hard":
                await _send_blocked_notice(update, str(user.id))
                return
        correction = correct_command_typo(text)
        if correction:
            await update.message.reply_text(
                f"💡 Did you mean `{correction}`?",
                parse_mode=ParseMode.MARKDOWN,
            )
        else:
            await update.message.reply_text(
                f"❓ Unknown command. Type `/help` to see available commands.",
                parse_mode=ParseMode.MARKDOWN,
            )
        return

    # Treat as inventory search
    logger.info(f"Message from {update.effective_user.first_name} [{update.effective_chat.id}]: {text}")

    # Log for AI summarization if in whitelisted group
    chat = update.effective_chat
    user = update.effective_user
    if chat and chat.type in ("group", "supergroup") and is_whitelisted_group(chat.id):
        log_group_message(chat.id, user.id, user.full_name, text)

    await _do_search(update, text)


# ──────────────────────────────────────────────────────────────────────────────
# Background Scheduler & Automated Reports
# ──────────────────────────────────────────────────────────────────────────────

async def send_automated_report(report_type: str):
    """Generate and send an automated report to all whitelisted groups."""
    groups = list_allowed_groups()
    if not groups:
        logger.info(f"No whitelisted groups found for {report_type} report.")
        return

    logger.info(f"Generating automated {report_type} report...")
    
    # 1. Refresh data first to ensure accuracy
    refresh_all_data()
    
    lines = []
    if report_type == "daily_inventory":
        # Simplified version of cmd_summary for executive overview
        inv = store.inventory
        total_qty = int(sum(r["in_stock"] for r in inv))
        unique_items = len(set(r["item_no"] for r in inv))
        
        # Low stock items
        low_stock = []
        by_item = group_inventory_by_item(inv)
        for k, v in by_item.items():
            if v["total"] < 500: # Standard threshold
                low_stock.append(v)
        
        lines = [
            "☀️ *Daily Inventory Report*",
            f"Total Stock: {total_qty:,} units",
            f"Unique Items: {unique_items:,}",
            "",
            f"⚠️ *Low Stock Alerts (<500kg):* {len(low_stock)} items",
        ]
        if low_stock:
            for item in sorted(low_stock, key=lambda x: x["total"])[:5]:
                lines.append(f"• {item['desc']}: {int(item['total']):,} units")
        
        lines.append("\n" + inv_source_footer())

    elif report_type == "weekly_collections":
        # Overdue AR reminder
        if not store.ar_rows:
            return
            
        overdue = [r for r in store.ar_rows if r["days_due"] >= 30]
        total_overdue = sum(r["balance"] for r in overdue)
        
        by_area = {}
        for r in overdue:
            a = r.get("area", "").strip() or "UNCLASSIFIED"
            by_area[a] = by_area.get(a, 0.0) + r["balance"]
            
        lines = [
            "💰 *Weekly Collection Reminder*",
            f"Total Overdue (30d+): {fmt_peso(total_overdue)}",
            "",
            "*Overdue by Area:*",
        ]
        for area, bal in sorted(by_area.items(), key=lambda x: x[1], reverse=True)[:10]:
            lines.append(f"• {area}: {fmt_peso(bal)}")
            
        lines.append("\n" + ar_source_footer())

    elif report_type == "weekly_slowmoving":
        # Slow moving items (30 days)
        inv = store.inventory
        by_item = group_inventory_by_item(inv)
        items_with_stock = {k: v for k, v in by_item.items() if v["total"] > 0}
        
        try:
            conn = _get_portal_conn()
            with conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT itemDescription AS product, SUM(CAST(quantity AS DECIMAL(12,3))) AS total_sold "
                        "FROM sales_transactions WHERE postingDate >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) GROUP BY itemDescription"
                    )
                    sales_rows = cur.fetchall()
            sales_map = {r["product"].strip().lower(): float(r["total_sold"]) for r in sales_rows if r["product"]}
            
            slow = []
            for item_no, data in items_with_stock.items():
                if sales_map.get(data["desc"].strip().lower(), 0.0) == 0:
                    slow.append(data)
            
            lines = [
                "🐢 *Weekly Slow-Moving Spotlight*",
                "_Items with stock but ZERO sales in 30 days._",
                "",
            ]
            for item in sorted(slow, key=lambda x: x["total"], reverse=True)[:10]:
                lines.append(f"• {item['desc']}: {int(item['total']):,} units")
            
            lines.append("\n" + inv_source_footer())
        except:
            return

    if not lines:
        return

    text = "\n".join(lines)
    for g in groups:
        chat_id = g["telegram_chat_id"]
        try:
            await tg_app.bot.send_message(chat_id, text, parse_mode=ParseMode.MARKDOWN)
            logger.info(f"Sent {report_type} report to {chat_id}")
        except Exception as e:
            logger.warning(f"Failed to send automated report to {chat_id}: {e}")


def background_scheduler_loop():
    """Background thread to trigger automated reports at specific times (PHT)."""
    import time
    logger.info("Background scheduler started.")
    
    # Track last sent date to avoid double sending
    last_daily = None
    last_weekly = None
    
    while True:
        try:
            now = datetime.now(PHT)
            today_str = now.strftime("%Y-%m-%d")
            week_str = now.strftime("%Y-%U") # Year-WeekNumber
            
            # 1. Daily Inventory Report: 8:00 AM PHT
            if now.hour == 8 and last_daily != today_str:
                asyncio.run_coroutine_threadsafe(send_automated_report("daily_inventory"), _loop)
                last_daily = today_str
            
            # 2. Weekly Reports (Monday 9:00 AM PHT)
            if now.weekday() == 0 and now.hour == 9 and last_weekly != week_str:
                asyncio.run_coroutine_threadsafe(send_automated_report("weekly_collections"), _loop)
                asyncio.run_coroutine_threadsafe(send_automated_report("weekly_slowmoving"), _loop)
                last_weekly = week_str
                
        except Exception as e:
            logger.error(f"Scheduler error: {e}")
            
        time.sleep(60) # Check every minute


def background_refresh_loop():
    import time
    while True:
        time.sleep(REFRESH_INTERVAL_SECONDS)
        logger.info("Auto-refresh triggered...")
        refresh_all_data()


# ──────────────────────────────────────────────────────────────────────────────
# Flask app + webhook
# ──────────────────────────────────────────────────────────────────────────────
flask_app = Flask(__name__)

tg_app = (
    Application.builder()
    .token(BOT_TOKEN)
    .build()
)

# Register all handlers
tg_app.add_handler(CommandHandler("start",      cmd_start))
tg_app.add_handler(CommandHandler("help",       cmd_help))
tg_app.add_handler(CommandHandler("myid",       cmd_myid))
tg_app.add_handler(CommandHandler("refresh",    cmd_refresh))

# Admin / Access Control
tg_app.add_handler(CommandHandler("register",   cmd_register))
tg_app.add_handler(CommandHandler("unregister", cmd_unregister))
tg_app.add_handler(CommandHandler("listusers",  cmd_listusers))
tg_app.add_handler(CommandHandler("testreports", cmd_testreports))
tg_app.add_handler(CommandHandler("summarize",  cmd_summarize))
tg_app.add_handler(CommandHandler("seen",       cmd_seen))
tg_app.add_handler(CommandHandler("accessmode", cmd_accessmode))

# Group whitelist (open chats for all members)
tg_app.add_handler(CommandHandler("whitelistgroup", cmd_whitelistgroup))
tg_app.add_handler(CommandHandler("unallowgroup", cmd_unallowgroup))
tg_app.add_handler(CommandHandler("allowedgroups", cmd_allowedgroups))
tg_app.add_handler(CommandHandler("seengroups", cmd_seengroups))

# Inventory
tg_app.add_handler(CommandHandler("summary",    cmd_summary))
tg_app.add_handler(CommandHandler("search",     cmd_search))
tg_app.add_handler(CommandHandler("check",      cmd_check))
tg_app.add_handler(CommandHandler("category",   cmd_category))
tg_app.add_handler(CommandHandler("expiring",   cmd_expiring))
tg_app.add_handler(CommandHandler("low",        cmd_low))
tg_app.add_handler(CommandHandler("lowstock",   cmd_lowstock))
tg_app.add_handler(CommandHandler("warehouse",  cmd_warehouse))
tg_app.add_handler(CommandHandler("slowmoving", cmd_slowmoving))
tg_app.add_handler(CallbackQueryHandler(cb_slowmoving_show_all, pattern=r"^slow_all\|"))
tg_app.add_handler(CallbackQueryHandler(cb_warehouse_show_all, pattern=r"^whs_all\|"))
tg_app.add_handler(CallbackQueryHandler(cb_expiring_show_all, pattern=r"^exp_all\|"))
tg_app.add_handler(CallbackQueryHandler(cb_check_show_all, pattern=r"^check_all\|"))
tg_app.add_handler(CallbackQueryHandler(cb_search_item_check, pattern=r"^check\|"))
tg_app.add_handler(CommandHandler("top",        cmd_top))
tg_app.add_handler(CommandHandler("components", cmd_components))
tg_app.add_handler(CommandHandler("components_expiring", cmd_components_expiring))
tg_app.add_handler(CallbackQueryHandler(cb_components_show_all, pattern=r"^comp_all\|"))

# AR
tg_app.add_handler(CommandHandler("arsummary",  cmd_ar_summary))
tg_app.add_handler(CommandHandler("ar",         cmd_ar))
tg_app.add_handler(CommandHandler("client",     cmd_client))
tg_app.add_handler(CommandHandler("aging",      cmd_aging))
tg_app.add_handler(CommandHandler("overdue",       cmd_overdue))
tg_app.add_handler(CommandHandler("top30overdue",  cmd_top30overdue))
tg_app.add_handler(CommandHandler("od30",          cmd_top30overdue))
tg_app.add_handler(CommandHandler("area",       cmd_area))
tg_app.add_handler(CommandHandler("agent",      cmd_agent))
tg_app.add_handler(CommandHandler("arsearch",   cmd_arsearch))
tg_app.add_handler(CommandHandler("arrefresh",  cmd_arrefresh))

# AP
tg_app.add_handler(CommandHandler("apsummary",  cmd_apsummary))
tg_app.add_handler(CommandHandler("ap_summary", cmd_ap_summary))
tg_app.add_handler(CommandHandler("ap",         cmd_ap))
tg_app.add_handler(CommandHandler("vendor",     cmd_vendor))
tg_app.add_handler(CommandHandler("apaging",    cmd_apaging))
tg_app.add_handler(CommandHandler("apoverdue",  cmd_apoverdue))
tg_app.add_handler(CommandHandler("aptop",      cmd_aptop))
tg_app.add_handler(CommandHandler("due_today",  cmd_due_today))
tg_app.add_handler(CommandHandler("aprefresh",  cmd_aprefresh))

# Sales Performance (Portal DB)
tg_app.add_handler(CallbackQueryHandler(cb_sales_area_pick, pattern=r"^sales_area\|"))
tg_app.add_handler(CommandHandler("sales",        cmd_sales))
tg_app.add_handler(CommandHandler("salesagent",   cmd_salesagent))
tg_app.add_handler(CommandHandler("salestarget",  cmd_salestarget))
tg_app.add_handler(CommandHandler("salesproduct", cmd_salesproduct))
tg_app.add_handler(CommandHandler("salesmonth",   cmd_salesmonth))

tg_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_free_text))


# Persistent event loop for all async operations
_loop: asyncio.AbstractEventLoop | None = None


async def _handle_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Global error handler — logs all unhandled exceptions from handlers."""
    logger.error(f"Unhandled exception in handler: {context.error}", exc_info=context.error)


@flask_app.route("/webhook", methods=["POST"])
def webhook():
    global _loop
    data = flask_request.get_json(force=True)
    update = Update.de_json(data, tg_app.bot)
    if _loop and not _loop.is_closed():
        future = asyncio.run_coroutine_threadsafe(tg_app.process_update(update), _loop)
        try:
            future.result(timeout=60)
        except Exception as e:
            logger.error(f"Webhook handler error: {e}")
    return "OK", 200


@flask_app.route("/", methods=["GET"])
def health():
    ts = store.last_refresh.strftime("%m/%d/%Y %I:%M %p PHT") if store.last_refresh else "not yet"
    inv = len(set(r["item_no"] for r in store.inventory))
    inv_src = store.inventory_source_ts or "unknown"
    ar_src = store.ar_source_ts or "unknown"
    ap_src = store.ap_source_ts or "unknown"
    return (
        f"Belcris Inventory Bot v3.0 — OK\n"
        f"Last refresh: {ts}\n"
        f"Items: {inv}\n"
        f"Inventory source: {inv_src} PHT\n"
        f"AR source: {ar_src} PHT\n"
        f"AP source: {ap_src} PHT"
    ), 200


async def setup_webhook():
    webhook_url = f"https://{WEBHOOK_HOST}/webhook"
    await tg_app.bot.set_webhook(url=webhook_url)
    logger.info(f"Webhook set to: {webhook_url}")


def _run_event_loop(loop: asyncio.AbstractEventLoop):
    """Run the event loop forever in a background thread."""
    asyncio.set_event_loop(loop)
    loop.run_forever()


def main():
    global _loop

    # Initial data load
    refresh_all_data()

    # Start background refresh thread
    t = threading.Thread(target=background_refresh_loop, daemon=True)
    t.start()

    # Start background scheduler thread
    ts = threading.Thread(target=background_scheduler_loop, daemon=True)
    ts.start()

    # Create a persistent event loop and run it in a background thread
    _loop = asyncio.new_event_loop()
    loop_thread = threading.Thread(target=_run_event_loop, args=(_loop,), daemon=True)
    loop_thread.start()

    # Initialize the telegram app and set webhook using the persistent loop
    future = asyncio.run_coroutine_threadsafe(tg_app.initialize(), _loop)
    future.result(timeout=30)
    tg_app.add_error_handler(_handle_error)
    if WEBHOOK_HOST:
        future2 = asyncio.run_coroutine_threadsafe(setup_webhook(), _loop)
        future2.result(timeout=30)

    # Start Flask (blocks main thread)
    logger.info(f"Starting Flask on port {PORT}...")
    flask_app.run(host="0.0.0.0", port=PORT)


if __name__ == "__main__":
    main()
